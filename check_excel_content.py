import sys
import os
from openpyxl import load_workbook

# Add current dir to path
sys.path.append(os.getcwd())

f = "backend/app/services/plantilla_valoracion.xlsx"

if os.path.exists(f):
    print(f"--- CHECKING CONTENT OF {f} ---")
    try:
        wb = load_workbook(f)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"Scanning sheet: {sheet}")
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if "ACTA" in cell.value or "0531" in cell.value:
                            print(f"FOUND in {sheet}!{cell.coordinate}: {cell.value}")
    except Exception as e:
        print(f"Error: {e}")
else:
    print(f"File not found: {f}")
