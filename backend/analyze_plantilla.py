from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Cargar plantilla
plantilla = load_workbook('app/services/plantilla_valoracion.xlsx')
ws = plantilla.active

print('=== ESTRUCTURA DE LA PLANTILLA ===\n')

# Mostrar celdas con texto (etiquetas) en cada sección importante
print('=== SECCIÓN IDENTIFICACIÓN (Filas 7-22) ===')
for row in range(7, 23):
    for col in range(1, 27):
        cell = ws.cell(row, col)
        if cell.value and str(cell.value).strip() != '':
            col_letter = get_column_letter(col)
            # Mostrar solo si tiene más de 2 caracteres (para filtrar marcas simples)
            if len(str(cell.value).strip()) > 2:
                print(f'  {col_letter}{row}: "{cell.value}"')

print('\n=== SECCIÓN INFORMACIÓN LABORAL (Filas 26-45) ===')
for row in range(26, 46):
    for col in range(1, 27):
        cell = ws.cell(row, col)
        if cell.value and str(cell.value).strip() != '':
            col_letter = get_column_letter(col)
            if len(str(cell.value).strip()) > 2:
                print(f'  {col_letter}{row}: "{cell.value}"')

print('\n=== SECCIÓN ACTIVIDAD LABORAL (Filas 55-59) ===')
for row in range(55, 60):
    row_data = []
    for col in range(1, 27):
        cell = ws.cell(row, col)
        if cell.value and str(cell.value).strip() != '':
            col_letter = get_column_letter(col)
            row_data.append(f'{col_letter}: "{cell.value}"')
    if row_data:
        print(f'  Fila {row}: {", ".join(row_data)}')

# Análisis de celdas merged (combinadas)
print('\n=== CELDAS COMBINADAS (MERGED) ===')
merged_ranges = ws.merged_cells.ranges
for merged_range in list(merged_ranges)[:30]:  # Primeras 30
    print(f'  {merged_range}')
