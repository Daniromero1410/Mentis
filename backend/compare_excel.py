from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Cargar ambos archivos
plantilla = load_workbook('app/services/plantilla_valoracion.xlsx')
generado = load_workbook('app/services/valoracion_Daniel_Romero__CC_1193581322_20260124_154632.xlsx')

ws_plantilla = plantilla.active
ws_generado = generado.active

print('=== COMPARACIÓN PLANTILLA vs GENERADO ===\n')
print('Mostrando celdas donde hay datos en el generado:\n')

# Recorrer todas las celdas hasta la fila 120
diferencias = []
for row in range(1, 121):
    for col in range(1, 27):  # A hasta Z
        cell_plantilla = ws_plantilla.cell(row, col)
        cell_generado = ws_generado.cell(row, col)

        # Si la celda generada tiene valor
        if cell_generado.value is not None and str(cell_generado.value).strip() != '':
            # Y es diferente de la plantilla
            if cell_plantilla.value != cell_generado.value:
                col_letter = get_column_letter(col)
                diferencias.append({
                    'celda': f'{col_letter}{row}',
                    'fila': row,
                    'plantilla': cell_plantilla.value,
                    'generado': cell_generado.value
                })

# Agrupar por filas para mejor visualización
print('DIFERENCIAS POR SECCIONES:\n')

# Identificación (filas 7-22)
print('=== IDENTIFICACIÓN (Filas 7-22) ===')
for diff in [d for d in diferencias if 7 <= d['fila'] <= 22]:
    print(f"  {diff['celda']}: \"{diff['generado']}\"")

# Info Laboral (filas 23-45)
print('\n=== INFORMACIÓN LABORAL (Filas 23-45) ===')
for diff in [d for d in diferencias if 23 <= d['fila'] <= 45]:
    print(f"  {diff['celda']}: \"{diff['generado']}\"")

# Historia Ocupacional (filas 46-53)
print('\n=== HISTORIA OCUPACIONAL (Filas 46-53) ===')
for diff in [d for d in diferencias if 46 <= d['fila'] <= 53]:
    print(f"  {diff['celda']}: \"{diff['generado']}\"")

# Actividad Laboral (filas 54-59)
print('\n=== ACTIVIDAD LABORAL ACTUAL (Filas 54-59) ===')
for diff in [d for d in diferencias if 54 <= d['fila'] <= 59]:
    print(f"  {diff['celda']}: \"{diff['generado']}\"")

# Factores de Riesgo (filas 60-110)
print('\n=== FACTORES DE RIESGO PSICOSOCIAL (Filas 60-110) ===')
riesgos = [d for d in diferencias if 60 <= d['fila'] <= 110]
print(f'  Total de evaluaciones marcadas: {len(riesgos)}')
if len(riesgos) > 0:
    print(f'  Primeras 10:')
    for diff in riesgos[:10]:
        print(f"    {diff['celda']}: \"{diff['generado']}\"")

# Concepto (filas 111-118)
print('\n=== CONCEPTO Y FIRMAS (Filas 111-118) ===')
for diff in [d for d in diferencias if 111 <= d['fila'] <= 118]:
    print(f"  {diff['celda']}: \"{diff['generado']}\"")

print(f'\n\nTOTAL DE DIFERENCIAS: {len(diferencias)}')
