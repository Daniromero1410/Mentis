from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

plantilla = load_workbook('app/services/plantilla_valoracion.xlsx')
ws = plantilla.active

print('=== ESTRUCTURA DE CHECKBOXES PARA NIVEL EDUCATIVO ===\n')

# Analizar fila por fila
for row in [16, 17, 18]:
    print(f'\n=== FILA {row} ===')

    # Obtener todas las celdas y sus valores
    celdas_info = []
    for col in range(1, 27):  # A-Z
        cell = ws.cell(row, col)
        col_letter = get_column_letter(col)

        # Verificar si es merged
        is_merged = False
        merged_parent = None
        for merged_range in ws.merged_cells.ranges:
            if f'{col_letter}{row}' in merged_range:
                is_merged = True
                merged_parent = f'{get_column_letter(merged_range.min_col)}{merged_range.min_row}'
                break

        if is_merged and merged_parent == f'{col_letter}{row}':
            # Es la celda principal de un merge
            celdas_info.append({
                'col': col_letter,
                'type': 'MERGED_PARENT',
                'value': cell.value,
                'merged_range': str([r for r in ws.merged_cells.ranges if f'{col_letter}{row}' in r][0])
            })
        elif not is_merged and cell.value:
            # Celda normal con valor
            celdas_info.append({
                'col': col_letter,
                'type': 'NORMAL',
                'value': cell.value
            })

    # Mostrar información
    for info in celdas_info:
        if info['type'] == 'MERGED_PARENT':
            print(f'  {info["col"]}{row}: "{info["value"]}" [{info["merged_range"]}]')
        else:
            print(f'  {info["col"]}{row}: "{info["value"]}" [celda normal]')

# Analizar estado civil para entender el patrón
print('\n\n=== ESTADO CIVIL (Fila 15) para referencia ===')
for col in range(1, 27):
    cell = ws.cell(15, col)
    col_letter = get_column_letter(col)
    if cell.value:
        is_merged = False
        for merged_range in ws.merged_cells.ranges:
            if f'{col_letter}15' in merged_range:
                is_merged = True
                break
        print(f'  {col_letter}15: "{cell.value}" {"[merged]" if is_merged else "[normal]"}')
