"""
Script para crear una plantilla Excel simplificada y fácil de mapear
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def crear_plantilla_simplificada():
    wb = Workbook()
    ws = wb.active
    ws.title = "Valoración"

    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    titulo_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')

    seccion_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    seccion_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    label_font = Font(name='Arial', size=10, bold=True)
    label_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    data_font = Font(name='Arial', size=10)

    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Configurar anchos de columnas
    ws.column_dimensions['A'].width = 25
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 15

    row = 1

    # ===== ENCABEZADO =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = 'FORMATO VALORACIÓN OCUPACIONAL'
    cell.font = titulo_font
    cell.fill = titulo_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 1

    # Fecha de valoración
    ws[f'A{row}'] = 'Fecha de valoración:'
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'] = ''  # Día
    ws[f'C{row}'] = ''  # Mes
    ws[f'D{row}'] = ''  # Año
    row += 2

    # ===== SECCIÓN 1: IDENTIFICACIÓN =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '1. IDENTIFICACIÓN DEL TRABAJADOR'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # Nombre
    ws[f'A{row}'] = 'Nombre completo:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    # Documento
    ws[f'A{row}'] = 'Documento:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    # Fecha de nacimiento
    ws[f'A{row}'] = 'Fecha de nacimiento:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws[f'B{row}'] = ''  # Día
    ws[f'C{row}'] = ''  # Mes
    ws[f'D{row}'] = ''  # Año
    ws[f'F{row}'] = 'Edad:'
    ws[f'F{row}'].font = label_font
    ws[f'G{row}'].border = border_thin
    row += 1

    # Sexo
    ws[f'A{row}'] = 'Sexo:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws[f'B{row}'] = 'M [ ]'
    ws[f'C{row}'] = 'F [ ]'
    row += 1

    # Estado civil
    ws[f'A{row}'] = 'Estado civil:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws[f'B{row}'] = 'Casado [ ]'
    ws[f'C{row}'] = 'Soltero [ ]'
    ws[f'D{row}'] = 'Unión libre [ ]'
    ws[f'E{row}'] = 'Separado [ ]'
    ws[f'F{row}'] = 'Viudo [ ]'
    row += 1

    # Nivel educativo
    ws[f'A{row}'] = 'Nivel educativo:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 1

    # Formación específica
    ws[f'A{row}'] = 'Formación específica:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    # Teléfonos
    ws[f'A{row}'] = 'Teléfonos:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    # Dirección
    ws[f'A{row}'] = 'Dirección:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    # Zona
    ws[f'A{row}'] = 'Zona:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws[f'B{row}'] = 'Urbana [ ]'
    ws[f'C{row}'] = 'Rural [ ]'
    row += 1

    # Ciudad, Departamento, País
    ws[f'A{row}'] = 'Ciudad:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws[f'B{row}'].border = border_thin
    ws[f'C{row}'] = 'Departamento:'
    ws[f'C{row}'].font = label_font
    ws[f'D{row}'].border = border_thin
    ws[f'E{row}'] = 'País:'
    ws[f'E{row}'].font = label_font
    ws[f'F{row}'].border = border_thin
    row += 2

    # ===== SECCIÓN 2: INFO LABORAL =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '2. INFORMACIÓN LABORAL'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # Empresa
    ws[f'A{row}'] = 'Empresa:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    # NIT
    ws[f'A{row}'] = 'NIT:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    # Ciudad empresa
    ws[f'A{row}'] = 'Ciudad:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws[f'B{row}'].border = border_thin
    ws[f'C{row}'] = 'Departamento:'
    ws[f'C{row}'].font = label_font
    ws[f'D{row}'].border = border_thin
    row += 1

    # Área
    ws[f'A{row}'] = 'Área:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    # Cargo
    ws[f'A{row}'] = 'Cargo:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    # Antiguedad
    ws[f'A{row}'] = 'Antigüedad en el cargo:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws[f'B{row}'].border = border_thin
    ws[f'C{row}'] = 'Antigüedad empresa:'
    ws[f'C{row}'].font = label_font
    ws[f'D{row}'].border = border_thin
    row += 1

    # Tipo de contrato
    ws[f'A{row}'] = 'Tipo de contrato:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 2

    # ===== SECCIÓN 3: HISTORIA OCUPACIONAL =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '3. HISTORIA OCUPACIONAL'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # Encabezados tabla
    ws[f'A{row}'] = 'Empresa'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws[f'A{row}'].border = border_thin
    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'] = 'Cargo/Funciones'
    ws[f'B{row}'].font = label_font
    ws[f'B{row}'].fill = label_fill
    ws[f'B{row}'].border = border_thin
    ws.merge_cells(f'E{row}:F{row}')
    ws[f'E{row}'] = 'Duración'
    ws[f'E{row}'].font = label_font
    ws[f'E{row}'].fill = label_fill
    ws[f'E{row}'].border = border_thin
    ws.merge_cells(f'G{row}:I{row}')
    ws[f'G{row}'] = 'Motivo de retiro'
    ws[f'G{row}'].font = label_font
    ws[f'G{row}'].fill = label_fill
    ws[f'G{row}'].border = border_thin
    row += 1

    # Filas de datos (3 empleos anteriores)
    for i in range(3):
        ws[f'A{row}'].border = border_thin
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'].border = border_thin
        ws.merge_cells(f'E{row}:F{row}')
        ws[f'E{row}'].border = border_thin
        ws.merge_cells(f'G{row}:I{row}')
        ws[f'G{row}'].border = border_thin
        row += 1

    row += 1

    # Otros oficios
    ws[f'A{row}'] = 'Otros oficios u ocupaciones:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 1

    # Oficios de interés
    ws[f'A{row}'] = 'Oficios de interés:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 2

    # ===== SECCIÓN 4: ACTIVIDAD LABORAL ACTUAL =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '4. ACTIVIDAD LABORAL ACTUAL'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # Nombre del cargo
    ws[f'A{row}'] = 'Nombre del cargo:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    # Tareas
    ws[f'A{row}'] = 'Tareas que realiza:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 1

    # Herramientas
    ws[f'A{row}'] = 'Herramientas/equipos:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 1

    # Horario
    ws[f'A{row}'] = 'Horario de trabajo:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    # EPP
    ws[f'A{row}'] = 'Elementos de protección:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    row += 2

    # ===== SECCIÓN 5: FACTORES DE RIESGO =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '5. FACTORES DE RIESGO PSICOSOCIALES'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # Encabezados
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = 'Factor de Riesgo'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws[f'A{row}'].border = border_thin
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    ws[f'G{row}'] = 'SI'
    ws[f'G{row}'].font = label_font
    ws[f'G{row}'].fill = label_fill
    ws[f'G{row}'].border = border_thin
    ws[f'G{row}'].alignment = Alignment(horizontal='center')

    ws[f'H{row}'] = 'NO'
    ws[f'H{row}'].font = label_font
    ws[f'H{row}'].fill = label_fill
    ws[f'H{row}'].border = border_thin
    ws[f'H{row}'].alignment = Alignment(horizontal='center')

    ws[f'I{row}'] = 'N/A'
    ws[f'I{row}'].font = label_font
    ws[f'I{row}'].fill = label_fill
    ws[f'I{row}'].border = border_thin
    ws[f'I{row}'].alignment = Alignment(horizontal='center')
    row += 1

    # Factores (categorías y items)
    factores = [
        ('DEMANDAS CUANTITATIVAS', 4),
        ('DEMANDAS DE CARGA MENTAL', 8),
        ('DEMANDAS EMOCIONALES', 6),
        ('EXIGENCIAS DE RESPONSABILIDAD', 6),
        ('CONSISTENCIA DE ROL', 7),
        ('DEMANDAS AMBIENTALES Y ESFUERZO FÍSICO', 14),
        ('DEMANDAS DE JORNADA', 4),
    ]

    for categoria, num_items in factores:
        # Categoría
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'] = categoria
        ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        ws[f'A{row}'].border = border_thin
        row += 1

        # Items
        for item_num in range(1, num_items + 1):
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = f'   Item {item_num}'
            ws[f'A{row}'].border = border_thin
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws[f'G{row}'].border = border_thin
            ws[f'H{row}'].border = border_thin
            ws[f'I{row}'].border = border_thin
            ws.row_dimensions[row].height = 20
            row += 1

    row += 1

    # ===== SECCIÓN 6: CONCEPTO FINAL =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '6. CONCEPTO FINAL'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # Apariencia personal
    ws[f'A{row}'] = 'Apariencia personal:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 1

    # Actitud
    ws[f'A{row}'] = 'Actitud:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 1

    # Concepto
    ws[f'A{row}'] = 'Concepto:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 60
    row += 1

    # Recomendaciones
    ws[f'A{row}'] = 'Recomendaciones:'
    ws[f'A{row}'].font = label_font
    ws[f'A{row}'].fill = label_fill
    ws.merge_cells(f'B{row}:I{row}')
    ws[f'B{row}'].border = border_thin
    ws[f'B{row}'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 60
    row += 2

    # Firma
    ws[f'A{row}'] = 'Nombre del evaluador:'
    ws[f'A{row}'].font = label_font
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    ws[f'A{row}'] = 'Firma:'
    ws[f'A{row}'].font = label_font
    ws.merge_cells(f'B{row}:E{row}')
    ws[f'B{row}'].border = border_thin
    row += 1

    ws[f'A{row}'] = 'Fecha:'
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'].border = border_thin
    ws[f'C{row}'].border = border_thin
    ws[f'D{row}'].border = border_thin

    # Guardar
    wb.save('app/services/plantilla_valoracion_simple.xlsx')
    print('✅ Plantilla simplificada creada exitosamente')
    print(f'📄 Total de filas utilizadas: {row}')
    print(f'📊 Celdas combinadas: aproximadamente 80-100 (vs 439 en la original)')
    print('\nArchivo guardado en: app/services/plantilla_valoracion_simple.xlsx')

if __name__ == '__main__':
    crear_plantilla_simplificada()
