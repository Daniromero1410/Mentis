from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

plantilla = load_workbook('app/services/plantilla_valoracion.xlsx')
generado = load_workbook('app/services/valoracion_Daniel_Romero__CC_1193581322_20260124_154632.xlsx')

ws_p = plantilla.active
ws_g = generado.active

print('=== ANÁLISIS DETALLADO DE PROBLEMAS DE MAPEO ===\n')

# Problema 1: Nivel Educativo
print('PROBLEMA 1: NIVEL EDUCATIVO (Filas 16-18)')
print('En la plantilla, las opciones están en:')
print('  Fila 16: H16="Formación empírica", N16="Básica primaria", T16="Bachillerato vocacional"')
print('  Fila 17: H17="Bachillerato modalidad", N17="Técnico/Tecnológico", T17="Universitario"')
print('  Fila 18: H18="Especialización/postgrado", N18="Formación informal", T18="Analfabeta"')
print('\nEn el archivo generado:')
for row in [16, 17, 18]:
    valores = []
    for col_letter in ['H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']:
        from openpyxl.utils import column_index_from_string
        col = column_index_from_string(col_letter)
        val = ws_g.cell(row, col).value
        if val and str(val).strip() != '':
            valores.append(f'{col_letter}{row}="{val}"')
    print(f'  Fila {row}: {", ".join(valores) if valores else "VACÍA - NO HAY MARCAS!"}')

# Problema 2: Actividad Laboral
print('\n\nPROBLEMA 2: ACTIVIDAD LABORAL ACTUAL (Filas 55-59)')
print('Las etiquetas en plantilla están en columna A:')
for row in range(55, 60):
    etiqueta_p = ws_p.cell(row, 1).value  # Columna A
    dato_g_col_a = ws_g.cell(row, 1).value  # Lo que hay en col A del generado
    dato_g_col_h = ws_g.cell(row, 8).value  # Lo que hay en col H del generado
    print(f'  Fila {row}:')
    print(f'    Plantilla A{row}: "{etiqueta_p}"')
    print(f'    Generado A{row}: "{dato_g_col_a}" <- SOBRESCRIBE LA ETIQUETA!')
    print(f'    Generado H{row}: "{dato_g_col_h}" <- DEBERÍA ESTAR AQUÍ')

# Problema 3: Documento
print('\n\nPROBLEMA 3: NÚMERO DE DOCUMENTO (Fila 11)')
print('En plantilla:')
print('  A11: "Número de documento" (etiqueta)')
print('En generado:')
print(f'  H11: "{ws_g.cell(11, 8).value}" <- Tiene TIPO + NÚMERO juntos')
print('DEBERÍA estar separado:')
print('  - Tipo de documento (CC/TI/CE) en una celda')
print('  - Número en otra celda')

# Verificar si hay celdas específicas para tipo de doc
print('\nBuscando celdas relacionadas con tipo de documento en fila 11:')
for col in range(1, 27):
    val_p = ws_p.cell(11, col).value
    val_g = ws_g.cell(11, col).value
    if val_p or val_g:
        col_letter = get_column_letter(col)
        print(f'  {col_letter}11: Plantilla="{val_p}", Generado="{val_g}"')

# Historia Ocupacional
print('\n\nPROBLEMA 4: HISTORIA OCUPACIONAL (Filas 49-51)')
print('En generado:')
for row in range(49, 52):
    print(f'  Fila {row}:')
    for col_letter in ['A', 'F', 'L', 'R']:
        from openpyxl.utils import column_index_from_string
        col = column_index_from_string(col_letter)
        val = ws_g.cell(row, col).value
        etiqueta = ws_p.cell(row, col).value
        if val:
            print(f'    {col_letter}{row}: "{val}" (etiqueta plantilla: "{etiqueta}")')
