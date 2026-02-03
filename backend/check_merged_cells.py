from openpyxl import load_workbook

plantilla = load_workbook('app/services/plantilla_valoracion.xlsx')
ws = plantilla.active

print('=== CELDAS COMBINADAS EN ÁREAS PROBLEMÁTICAS ===\n')

# Filas 55-59 (Actividad Laboral)
print('ACTIVIDAD LABORAL (Filas 55-59):')
for merged_range in ws.merged_cells.ranges:
    if any(merged_range.min_row <= row <= merged_range.max_row for row in range(55, 60)):
        print(f'  {merged_range}')

# Filas 16-18 (Nivel Educativo)
print('\nNIVEL EDUCATIVO (Filas 16-18):')
for merged_range in ws.merged_cells.ranges:
    if any(merged_range.min_row <= row <= merged_range.max_row for row in range(16, 19)):
        print(f'  {merged_range}')

# Fila 11 (Documento)
print('\nDOCUMENTO (Fila 11):')
for merged_range in ws.merged_cells.ranges:
    if merged_range.min_row == 11 or merged_range.max_row == 11:
        print(f'  {merged_range}')

# Verificar específicamente dónde se deben escribir los datos
from openpyxl.utils import get_column_letter

print('\n\n=== VERIFICACIÓN DE CELDAS ESPECÍFICAS ===')

# Actividad Laboral
print('\nActividad Laboral - Fila 55:')
for col in range(1, 27):
    cell = ws.cell(55, col)
    col_letter = get_column_letter(col)
    # Verificar si es parte de merged cell
    is_merged = False
    merged_parent = None
    for merged_range in ws.merged_cells.ranges:
        if f'{col_letter}55' in merged_range:
            is_merged = True
            merged_parent = f'{get_column_letter(merged_range.min_col)}{merged_range.min_row}'
            break

    if is_merged:
        print(f'  {col_letter}55: MERGED (parent: {merged_parent})')
    elif cell.value:
        print(f'  {col_letter}55: "{cell.value}"')

print('\nNivel Educativo - Fila 16:')
for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W']:
    from openpyxl.utils import column_index_from_string
    col_num = column_index_from_string(col)
    cell = ws.cell(16, col_num)

    # Verificar si es parte de merged cell
    is_merged = False
    merged_parent = None
    for merged_range in ws.merged_cells.ranges:
        if f'{col}16' in merged_range:
            is_merged = True
            merged_parent = f'{get_column_letter(merged_range.min_col)}{merged_range.min_row}'
            break

    if is_merged:
        print(f'  {col}16: MERGED (parent: {merged_parent})')
    elif cell.value and len(str(cell.value).strip()) > 1:
        print(f'  {col}16: "{cell.value}"')
