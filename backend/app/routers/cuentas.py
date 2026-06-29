"""
Router del módulo de Cuentas (relación de servicios mensuales).

- Terapeuta: registra/edita/cierra sus servicios del mes (sin precios).
- Administrador: ve el consolidado de todos, asigna precios (catálogo de
  tarifas) y consulta totales con retefuente (12%) y pago del 70%.
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlmodel import Session, select, func
from typing import Optional, List
from datetime import datetime
import os
import tempfile

from app.database.connection import get_session
from app.models.usuario import Usuario
from app.models.cuenta import (
    ServicioCuenta, CierreMensual, TarifaCuenta, CatalogoServicio, Notificacion, EstadoCierre
)
from app.schemas.cuenta import (
    ServicioCuentaCreate, ServicioCuentaUpdate,
    ServicioCuentaTerapeutaRead, ServicioCuentaAdminRead,
    PrecioUpdate, CierreRead, TarifaCreate, TarifaRead,
    ConsolidadoAdminResponse, TotalesPorArl,
    CatalogoServicioCreate, CatalogoServicioUpdate, CatalogoServicioRead,
)
from app.services.auth import get_current_user, get_current_admin

router = APIRouter(prefix="/cuentas", tags=["Cuentas"])

RETEFUENTE_FACTOR = 0.12    # retefuente = lo que se descuenta (12%)
PAGO_PARCIAL_FACTOR = 0.70  # 70% del valor posterior a retefuente


# ═════════════════════════════════════════════════════════════════════
# Helpers
# ═════════════════════════════════════════════════════════════════════
def _is_admin(user: Usuario) -> bool:
    rol = user.rol.value if hasattr(user.rol, "value") else str(user.rol)
    return rol.lower() == "admin"


def _resolver_cups(session: Session, servicio_nombre: Optional[str]) -> Optional[str]:
    """Devuelve el CUPS del catálogo para un nombre de servicio (o None)."""
    if not servicio_nombre:
        return None
    cat = session.exec(
        select(CatalogoServicio).where(
            func.lower(CatalogoServicio.nombre) == servicio_nombre.strip().lower()
        )
    ).first()
    return cat.cups if cat else None


def _get_or_create_cierre(session: Session, terapeuta_id: int, mes: int, anio: int) -> CierreMensual:
    cierre = session.exec(
        select(CierreMensual).where(
            CierreMensual.terapeuta_id == terapeuta_id,
            CierreMensual.periodo_mes == mes,
            CierreMensual.periodo_anio == anio,
        )
    ).first()
    if not cierre:
        cierre = CierreMensual(terapeuta_id=terapeuta_id, periodo_mes=mes, periodo_anio=anio,
                               estado=EstadoCierre.ABIERTO.value)
        session.add(cierre)
        session.commit()
        session.refresh(cierre)
    return cierre


def _periodo_esta_cerrado(session: Session, terapeuta_id: int, mes: int, anio: int) -> bool:
    cierre = session.exec(
        select(CierreMensual).where(
            CierreMensual.terapeuta_id == terapeuta_id,
            CierreMensual.periodo_mes == mes,
            CierreMensual.periodo_anio == anio,
        )
    ).first()
    return bool(cierre and cierre.estado in (EstadoCierre.CERRADO.value, EstadoCierre.REVISADO.value))


# ═════════════════════════════════════════════════════════════════════
# TERAPEUTA — Mis servicios
# ═════════════════════════════════════════════════════════════════════
@router.get("/mis-servicios", response_model=List[ServicioCuentaTerapeutaRead])
def listar_mis_servicios(
    mes: int = Query(...),
    anio: int = Query(...),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    servicios = session.exec(
        select(ServicioCuenta).where(
            ServicioCuenta.terapeuta_id == current_user.id,
            ServicioCuenta.periodo_mes == mes,
            ServicioCuenta.periodo_anio == anio,
        ).order_by(ServicioCuenta.id)
    ).all()
    return list(servicios)


@router.post("/mis-servicios", response_model=ServicioCuentaTerapeutaRead, status_code=201)
def crear_mi_servicio(
    data: ServicioCuentaCreate,
    mes: int = Query(...),
    anio: int = Query(...),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    if _periodo_esta_cerrado(session, current_user.id, mes, anio):
        raise HTTPException(400, "El mes está cerrado. No se pueden agregar servicios.")
    servicio = ServicioCuenta(
        terapeuta_id=current_user.id,
        periodo_mes=mes,
        periodo_anio=anio,
        **data.model_dump(),
    )
    # El CUPS siempre se toma del catálogo (autoritativo)
    servicio.cups = _resolver_cups(session, servicio.servicio)
    session.add(servicio)
    # Asegurar que exista el cierre (abierto) del periodo
    _get_or_create_cierre(session, current_user.id, mes, anio)
    session.commit()
    session.refresh(servicio)
    return servicio


@router.put("/mis-servicios/{servicio_id}", response_model=ServicioCuentaTerapeutaRead)
def actualizar_mi_servicio(
    servicio_id: int,
    data: ServicioCuentaUpdate,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    servicio = session.get(ServicioCuenta, servicio_id)
    if not servicio or servicio.terapeuta_id != current_user.id:
        raise HTTPException(404, "Servicio no encontrado")
    if _periodo_esta_cerrado(session, current_user.id, servicio.periodo_mes, servicio.periodo_anio):
        raise HTTPException(400, "El mes está cerrado. No se puede editar.")
    for k, v in data.model_dump().items():
        setattr(servicio, k, v)
    # El CUPS siempre se resuelve del catálogo según el servicio elegido
    servicio.cups = _resolver_cups(session, servicio.servicio)
    servicio.updated_at = datetime.utcnow()
    session.add(servicio)
    session.commit()
    session.refresh(servicio)
    return servicio


@router.delete("/mis-servicios/{servicio_id}", status_code=204)
def eliminar_mi_servicio(
    servicio_id: int,
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    servicio = session.get(ServicioCuenta, servicio_id)
    if not servicio or servicio.terapeuta_id != current_user.id:
        raise HTTPException(404, "Servicio no encontrado")
    if _periodo_esta_cerrado(session, current_user.id, servicio.periodo_mes, servicio.periodo_anio):
        raise HTTPException(400, "El mes está cerrado. No se puede eliminar.")
    session.delete(servicio)
    session.commit()
    return None


# ── Estado del cierre del terapeuta ──────────────────────────────────
@router.get("/mi-cierre", response_model=CierreRead)
def estado_mi_cierre(
    mes: int = Query(...),
    anio: int = Query(...),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    cierre = _get_or_create_cierre(session, current_user.id, mes, anio)
    total = session.exec(
        select(func.count()).select_from(ServicioCuenta).where(
            ServicioCuenta.terapeuta_id == current_user.id,
            ServicioCuenta.periodo_mes == mes,
            ServicioCuenta.periodo_anio == anio,
        )
    ).one()
    return CierreRead(
        id=cierre.id, terapeuta_id=cierre.terapeuta_id,
        terapeuta_nombre=f"{current_user.nombre} {current_user.apellido}",
        periodo_mes=cierre.periodo_mes, periodo_anio=cierre.periodo_anio,
        estado=cierre.estado, fecha_cierre=cierre.fecha_cierre, total_servicios=total,
    )


@router.post("/cerrar-mes", response_model=CierreRead)
def cerrar_mes(
    mes: int = Query(...),
    anio: int = Query(...),
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    cierre = _get_or_create_cierre(session, current_user.id, mes, anio)
    if cierre.estado in (EstadoCierre.CERRADO.value, EstadoCierre.REVISADO.value):
        raise HTTPException(400, "El mes ya está cerrado.")

    total = session.exec(
        select(func.count()).select_from(ServicioCuenta).where(
            ServicioCuenta.terapeuta_id == current_user.id,
            ServicioCuenta.periodo_mes == mes,
            ServicioCuenta.periodo_anio == anio,
        )
    ).one()
    if total == 0:
        raise HTTPException(400, "No hay servicios registrados en este mes.")

    cierre.estado = EstadoCierre.CERRADO.value
    cierre.fecha_cierre = datetime.utcnow()
    cierre.updated_at = datetime.utcnow()
    session.add(cierre)

    # Notificar a todos los administradores
    nombre_completo = f"{current_user.nombre} {current_user.apellido}"
    meses_es = ["", "enero", "febrero", "marzo", "abril", "mayo", "junio",
                "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    admins = session.exec(select(Usuario).where(Usuario.rol == "admin")).all()
    for admin in admins:
        session.add(Notificacion(
            usuario_destino_id=admin.id,
            tipo="cierre_mensual",
            titulo="Cierre de cuenta mensual",
            mensaje=f"{nombre_completo} cerró su relación de servicios de {meses_es[mes]} {anio} ({total} servicios).",
            link="/dashboard/cuentas",
        ))

    session.commit()
    session.refresh(cierre)
    return CierreRead(
        id=cierre.id, terapeuta_id=cierre.terapeuta_id,
        terapeuta_nombre=nombre_completo,
        periodo_mes=cierre.periodo_mes, periodo_anio=cierre.periodo_anio,
        estado=cierre.estado, fecha_cierre=cierre.fecha_cierre, total_servicios=total,
    )


# ═════════════════════════════════════════════════════════════════════
# ADMIN — Consolidado
# ═════════════════════════════════════════════════════════════════════
@router.get("/admin/servicios", response_model=ConsolidadoAdminResponse)
def consolidado_admin(
    mes: int = Query(...),
    anio: int = Query(...),
    arl: Optional[str] = None,
    terapeuta_id: Optional[int] = None,
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    q = select(ServicioCuenta).where(
        ServicioCuenta.periodo_mes == mes,
        ServicioCuenta.periodo_anio == anio,
    )
    if arl:
        q = q.where(ServicioCuenta.arl == arl)
    if terapeuta_id:
        q = q.where(ServicioCuenta.terapeuta_id == terapeuta_id)
    servicios = session.exec(q.order_by(ServicioCuenta.arl, ServicioCuenta.terapeuta_id, ServicioCuenta.id)).all()

    # Cache de nombres de terapeutas
    nombres: dict = {}
    for u in session.exec(select(Usuario)).all():
        nombres[u.id] = f"{u.nombre} {u.apellido}"

    # Servicios que permiten viáticos (según el catálogo)
    servicios_con_viaticos = {
        c.nombre.strip().lower()
        for c in session.exec(select(CatalogoServicio).where(CatalogoServicio.permite_viaticos == True)).all()  # noqa: E712
    }

    items: List[ServicioCuentaAdminRead] = []
    totales: dict = {}  # arl -> [count, bruto, viaticos]
    for s in servicios:
        precio = s.precio_unitario or 0
        cant = s.cantidad or 0
        total_linea = precio * cant
        viat = s.viaticos or 0
        permite_v = bool(s.servicio and s.servicio.strip().lower() in servicios_con_viaticos)
        items.append(ServicioCuentaAdminRead(
            id=s.id, terapeuta_id=s.terapeuta_id,
            terapeuta_nombre=nombres.get(s.terapeuta_id, ""),
            periodo_mes=s.periodo_mes, periodo_anio=s.periodo_anio,
            nombre_usuario=s.nombre_usuario, tipo_documento=s.tipo_documento,
            numero_documento=s.numero_documento, arl=s.arl, servicio=s.servicio, cups=s.cups,
            numero_autorizacion=s.numero_autorizacion,
            fecha_realizacion=s.fecha_realizacion, fecha_autorizacion=s.fecha_autorizacion,
            carpeta_cargue=s.carpeta_cargue, cantidad=s.cantidad,
            recomendaciones=s.recomendaciones, en_pleno=s.en_pleno, pcl=s.pcl,
            precio_unitario=s.precio_unitario, viaticos=s.viaticos, nota_admin=s.nota_admin,
            total=total_linea, permite_viaticos=permite_v,
        ))
        key = s.arl or "SIN ARL"
        if key not in totales:
            totales[key] = [0, 0.0, 0.0]
        totales[key][0] += 1
        totales[key][1] += total_linea
        totales[key][2] += viat

    totales_por_arl = []
    bruto_total = 0.0
    viaticos_total = 0.0
    for arl_name, (cnt, bruto, viat) in sorted(totales.items()):
        rete = round(bruto * RETEFUENTE_FACTOR)       # lo que se descuenta (12%)
        posterior = bruto - rete                       # lo que queda (88%)
        pago = round(posterior * PAGO_PARCIAL_FACTOR)  # 70% del valor posterior
        totales_por_arl.append(TotalesPorArl(
            arl=arl_name, total_servicios=cnt, valor_bruto=bruto,
            retefuente=rete, valor_posterior_retefuente=posterior,
            viaticos=viat, pago_70=pago, total_a_pagar=pago + viat,
        ))
        bruto_total += bruto
        viaticos_total += viat

    rete_total = round(bruto_total * RETEFUENTE_FACTOR)
    posterior_total = bruto_total - rete_total
    pago_total = round(posterior_total * PAGO_PARCIAL_FACTOR)

    return ConsolidadoAdminResponse(
        servicios=items,
        totales_por_arl=totales_por_arl,
        valor_bruto_total=bruto_total,
        retefuente_total=rete_total,
        valor_posterior_retefuente_total=posterior_total,
        viaticos_total=viaticos_total,
        pago_70_total=pago_total,
        total_a_pagar_total=pago_total + viaticos_total,
    )


# ── Admin: exportar consolidado a Excel ──────────────────────────────
@router.get("/admin/exportar", response_class=FileResponse)
def exportar_excel(
    mes: int = Query(...),
    anio: int = Query(...),
    arl: Optional[str] = None,
    terapeuta_id: Optional[int] = None,
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    MESES = ["", "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
             "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

    q = select(ServicioCuenta).where(
        ServicioCuenta.periodo_mes == mes,
        ServicioCuenta.periodo_anio == anio,
    )
    if arl:
        q = q.where(ServicioCuenta.arl == arl)
    if terapeuta_id:
        q = q.where(ServicioCuenta.terapeuta_id == terapeuta_id)
    servicios = session.exec(q.order_by(ServicioCuenta.arl, ServicioCuenta.terapeuta_id, ServicioCuenta.id)).all()

    nombres: dict = {u.id: f"{u.nombre} {u.apellido}" for u in session.exec(select(Usuario)).all()}

    # ── Estilos ──
    naranja = PatternFill(start_color="F07820", end_color="F07820", fill_type="solid")
    gris = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    blanco_bold = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    centro = Alignment(horizontal="center", vertical="center")
    borde = Border(*[Side(style="thin", color="BDBDBD")] * 4)

    def fmt_fecha(f):
        if not f:
            return ""
        return f.strftime("%d/%m/%Y") if hasattr(f, "strftime") else str(f)

    # Servicios que permiten viáticos
    servicios_con_viaticos = {
        c.nombre.strip().lower()
        for c in session.exec(select(CatalogoServicio).where(CatalogoServicio.permite_viaticos == True)).all()  # noqa: E712
    }

    def _sanitizar_hoja(nombre: str) -> str:
        """Nombre de hoja válido para Excel (máx 31 chars, sin caracteres prohibidos)."""
        for ch in '[]:*?/\\':
            nombre = nombre.replace(ch, ' ')
        nombre = nombre.strip() or "Hoja"
        return nombre[:31]

    def _escribir_hoja(ws, lista, titulo):
        """Escribe la tabla completa (servicios + totales) en una hoja."""
        ws.merge_cells("A1:P1")
        c = ws["A1"]
        c.value = titulo
        c.font = Font(bold=True, size=14)
        c.alignment = centro
        ws.row_dimensions[1].height = 24

        headers = ["Terapeuta", "Usuario", "Tipo Doc", "Documento", "ARL", "Servicio", "CUPS",
                   "Autorización", "F. Realización", "F. Autorización", "Carpeta",
                   "Cant.", "Precio Unit.", "Total", "Viáticos", "Total a pagar"]
        hrow = 3
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=hrow, column=col, value=h)
            cell.fill = naranja
            cell.font = blanco_bold
            cell.alignment = centro
            cell.border = borde

        r = hrow + 1
        totales: dict = {}  # arl -> [count, bruto, viaticos]
        for s in lista:
            precio = s.precio_unitario or 0
            cant = s.cantidad or 0
            total = precio * cant
            permite_v = bool(s.servicio and s.servicio.strip().lower() in servicios_con_viaticos)
            viat = (s.viaticos or 0) if permite_v else 0
            valores = [
                nombres.get(s.terapeuta_id, ""), s.nombre_usuario or "", s.tipo_documento or "",
                s.numero_documento or "", s.arl or "", s.servicio or "", s.cups or "", s.numero_autorizacion or "",
                fmt_fecha(s.fecha_realizacion), fmt_fecha(s.fecha_autorizacion), s.carpeta_cargue or "",
                cant, precio, total, (viat if permite_v else ""), total + viat,
            ]
            for col, val in enumerate(valores, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = borde
                if col in (13, 14, 15, 16):  # precio, total, viáticos, total a pagar
                    cell.number_format = '"$"#,##0'
                if col == 12:  # cantidad
                    cell.alignment = centro
            key = s.arl or "SIN ARL"
            if key not in totales:
                totales[key] = [0, 0.0, 0.0]
            totales[key][0] += 1
            totales[key][1] += total
            totales[key][2] += viat
            r += 1

        # Totales por ARL
        r += 1
        ws.cell(row=r, column=1, value="TOTALES POR ARL").font = bold
        r += 1
        th = ["ARL", "Servicios", "Valor bruto", "Retefuente (12%)", "Valor posterior retefuente", "Pago 70%", "Viáticos", "Total a pagar"]
        for col, h in enumerate(th, start=1):
            cell = ws.cell(row=r, column=col, value=h)
            cell.fill = gris
            cell.font = bold
            cell.alignment = centro
            cell.border = borde
        r += 1
        bruto_total = 0.0
        viaticos_total = 0.0
        for arl_name, (cnt, bruto, viat) in sorted(totales.items()):
            rete = round(bruto * 0.12)
            posterior = bruto - rete
            pago = round(posterior * 0.70)
            fila = [arl_name, cnt, bruto, rete, posterior, pago, viat, pago + viat]
            for col, val in enumerate(fila, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = borde
                if col in (3, 4, 5, 6, 7, 8):
                    cell.number_format = '"$"#,##0'
                if col == 2:
                    cell.alignment = centro
            bruto_total += bruto
            viaticos_total += viat
            r += 1

        # Gran total
        rete_t = round(bruto_total * 0.12)
        posterior_t = bruto_total - rete_t
        pago_t = round(posterior_t * 0.70)
        fila = ["TOTAL GENERAL", "", bruto_total, rete_t, posterior_t, pago_t, viaticos_total, pago_t + viaticos_total]
        for col, val in enumerate(fila, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = bold
            cell.fill = gris
            cell.border = borde
            if col in (3, 4, 5, 6, 7, 8):
                cell.number_format = '"$"#,##0'

        anchos = [22, 26, 10, 14, 16, 30, 12, 14, 14, 14, 16, 7, 14, 14, 14, 15]
        for idx_a, w in enumerate(anchos, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx_a)].width = w

    wb = openpyxl.Workbook()

    # Hoja principal: todos los servicios
    ws_main = wb.active
    ws_main.title = _sanitizar_hoja(f"Consolidado {MESES[mes][:3]} {anio}")
    _escribir_hoja(ws_main, servicios, f"CONSOLIDADO DE CUENTAS — {MESES[mes]} {anio}")

    # Una hoja por terapeuta (en orden de aparición)
    terapeutas_orden = []
    por_terapeuta: dict = {}
    for s in servicios:
        if s.terapeuta_id not in por_terapeuta:
            por_terapeuta[s.terapeuta_id] = []
            terapeutas_orden.append(s.terapeuta_id)
        por_terapeuta[s.terapeuta_id].append(s)

    usadas = {ws_main.title.lower()}
    for tid in terapeutas_orden:
        nombre_t = nombres.get(tid, f"Terapeuta {tid}")
        base = _sanitizar_hoja(nombre_t)
        titulo_hoja = base
        n = 2
        while titulo_hoja.lower() in usadas:
            titulo_hoja = _sanitizar_hoja(f"{base[:27]} ({n})")
            n += 1
        usadas.add(titulo_hoja.lower())
        ws_t = wb.create_sheet(title=titulo_hoja)
        _escribir_hoja(ws_t, por_terapeuta[tid], f"{nombre_t} — {MESES[mes]} {anio}")

    # Guardar a archivo temporal
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    wb.save(tmp.name)
    filename = f"Cuentas_{MESES[mes]}_{anio}.xlsx"
    return FileResponse(
        path=tmp.name,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.put("/admin/servicios/{servicio_id}/precio", response_model=ServicioCuentaAdminRead)
def asignar_precio(
    servicio_id: int,
    data: PrecioUpdate,
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    servicio = session.get(ServicioCuenta, servicio_id)
    if not servicio:
        raise HTTPException(404, "Servicio no encontrado")
    if data.precio_unitario is not None:
        servicio.precio_unitario = data.precio_unitario
    if data.viaticos is not None:
        servicio.viaticos = data.viaticos
    if data.nota_admin is not None:
        servicio.nota_admin = data.nota_admin
    servicio.updated_at = datetime.utcnow()
    session.add(servicio)
    session.commit()
    session.refresh(servicio)
    u = session.get(Usuario, servicio.terapeuta_id)
    cat = session.exec(
        select(CatalogoServicio).where(
            func.lower(CatalogoServicio.nombre) == (servicio.servicio or "").strip().lower()
        )
    ).first()
    return ServicioCuentaAdminRead(
        id=servicio.id, terapeuta_id=servicio.terapeuta_id,
        terapeuta_nombre=f"{u.nombre} {u.apellido}" if u else "",
        periodo_mes=servicio.periodo_mes, periodo_anio=servicio.periodo_anio,
        nombre_usuario=servicio.nombre_usuario, tipo_documento=servicio.tipo_documento,
        numero_documento=servicio.numero_documento, arl=servicio.arl, servicio=servicio.servicio, cups=servicio.cups,
        numero_autorizacion=servicio.numero_autorizacion,
        fecha_realizacion=servicio.fecha_realizacion, fecha_autorizacion=servicio.fecha_autorizacion,
        carpeta_cargue=servicio.carpeta_cargue, cantidad=servicio.cantidad,
        recomendaciones=servicio.recomendaciones, en_pleno=servicio.en_pleno, pcl=servicio.pcl,
        precio_unitario=servicio.precio_unitario, viaticos=servicio.viaticos, nota_admin=servicio.nota_admin,
        total=(servicio.precio_unitario or 0) * (servicio.cantidad or 0),
        permite_viaticos=bool(cat and cat.permite_viaticos),
    )


@router.post("/admin/aplicar-tarifas")
def aplicar_tarifas(
    mes: int = Query(...),
    anio: int = Query(...),
    solo_vacios: bool = Query(True),
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    """Aplica el catálogo de tarifas (ARL+servicio) a los servicios del periodo."""
    tarifas = session.exec(select(TarifaCuenta)).all()
    catalogo = {(t.arl, t.servicio): t.precio_unitario for t in tarifas}

    servicios = session.exec(
        select(ServicioCuenta).where(
            ServicioCuenta.periodo_mes == mes,
            ServicioCuenta.periodo_anio == anio,
        )
    ).all()
    actualizados = 0
    for s in servicios:
        if solo_vacios and s.precio_unitario:
            continue
        precio = catalogo.get((s.arl, s.servicio))
        if precio is not None:
            s.precio_unitario = precio
            s.updated_at = datetime.utcnow()
            session.add(s)
            actualizados += 1
    session.commit()
    return {"actualizados": actualizados, "total_servicios": len(servicios)}


# ── Admin: cierres por terapeuta ─────────────────────────────────────
@router.get("/admin/cierres", response_model=List[CierreRead])
def listar_cierres(
    mes: int = Query(...),
    anio: int = Query(...),
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    cierres = session.exec(
        select(CierreMensual).where(
            CierreMensual.periodo_mes == mes,
            CierreMensual.periodo_anio == anio,
        )
    ).all()
    out = []
    for c in cierres:
        u = session.get(Usuario, c.terapeuta_id)
        total = session.exec(
            select(func.count()).select_from(ServicioCuenta).where(
                ServicioCuenta.terapeuta_id == c.terapeuta_id,
                ServicioCuenta.periodo_mes == mes,
                ServicioCuenta.periodo_anio == anio,
            )
        ).one()
        out.append(CierreRead(
            id=c.id, terapeuta_id=c.terapeuta_id,
            terapeuta_nombre=f"{u.nombre} {u.apellido}" if u else "",
            periodo_mes=c.periodo_mes, periodo_anio=c.periodo_anio,
            estado=c.estado, fecha_cierre=c.fecha_cierre, total_servicios=total,
        ))
    return out


@router.put("/admin/cierres/{cierre_id}/estado", response_model=CierreRead)
def cambiar_estado_cierre(
    cierre_id: int,
    estado: str = Query(...),
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    if estado not in [e.value for e in EstadoCierre]:
        raise HTTPException(400, "Estado inválido")
    cierre = session.get(CierreMensual, cierre_id)
    if not cierre:
        raise HTTPException(404, "Cierre no encontrado")
    cierre.estado = estado
    if estado == EstadoCierre.ABIERTO.value:
        cierre.fecha_cierre = None
    cierre.updated_at = datetime.utcnow()
    session.add(cierre)
    session.commit()
    session.refresh(cierre)
    u = session.get(Usuario, cierre.terapeuta_id)
    return CierreRead(
        id=cierre.id, terapeuta_id=cierre.terapeuta_id,
        terapeuta_nombre=f"{u.nombre} {u.apellido}" if u else "",
        periodo_mes=cierre.periodo_mes, periodo_anio=cierre.periodo_anio,
        estado=cierre.estado, fecha_cierre=cierre.fecha_cierre,
    )


# ═════════════════════════════════════════════════════════════════════
# ADMIN — Catálogo de tarifas
# ═════════════════════════════════════════════════════════════════════
@router.get("/admin/tarifas", response_model=List[TarifaRead])
def listar_tarifas(
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    return list(session.exec(select(TarifaCuenta).order_by(TarifaCuenta.arl, TarifaCuenta.servicio)).all())


@router.post("/admin/tarifas", response_model=TarifaRead, status_code=201)
def crear_tarifa(
    data: TarifaCreate,
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    existente = session.exec(
        select(TarifaCuenta).where(
            TarifaCuenta.arl == data.arl, TarifaCuenta.servicio == data.servicio
        )
    ).first()
    if existente:
        existente.precio_unitario = data.precio_unitario
        existente.updated_at = datetime.utcnow()
        session.add(existente)
        session.commit()
        session.refresh(existente)
        return existente
    tarifa = TarifaCuenta(**data.model_dump())
    session.add(tarifa)
    session.commit()
    session.refresh(tarifa)
    return tarifa


@router.put("/admin/tarifas/{tarifa_id}", response_model=TarifaRead)
def actualizar_tarifa(
    tarifa_id: int,
    data: TarifaCreate,
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    tarifa = session.get(TarifaCuenta, tarifa_id)
    if not tarifa:
        raise HTTPException(404, "Tarifa no encontrada")
    tarifa.arl = data.arl
    tarifa.servicio = data.servicio
    tarifa.precio_unitario = data.precio_unitario
    tarifa.updated_at = datetime.utcnow()
    session.add(tarifa)
    session.commit()
    session.refresh(tarifa)
    return tarifa


@router.delete("/admin/tarifas/{tarifa_id}", status_code=204)
def eliminar_tarifa(
    tarifa_id: int,
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    tarifa = session.get(TarifaCuenta, tarifa_id)
    if not tarifa:
        raise HTTPException(404, "Tarifa no encontrada")
    session.delete(tarifa)
    session.commit()
    return None


# ═════════════════════════════════════════════════════════════════════
# CATÁLOGO DE SERVICIOS (lectura para todos, edición solo admin)
# ═════════════════════════════════════════════════════════════════════
@router.get("/servicios-catalogo", response_model=List[CatalogoServicioRead])
def listar_servicios_catalogo(
    session: Session = Depends(get_session),
    current_user: Usuario = Depends(get_current_user),
):
    """Servicios activos — disponibles para todos los usuarios."""
    items = session.exec(
        select(CatalogoServicio).where(CatalogoServicio.activo == True)  # noqa: E712
        .order_by(CatalogoServicio.orden, CatalogoServicio.nombre)
    ).all()
    return list(items)


@router.get("/admin/servicios-catalogo", response_model=List[CatalogoServicioRead])
def listar_servicios_catalogo_admin(
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    """Todos los servicios (activos e inactivos) para gestión del admin."""
    items = session.exec(
        select(CatalogoServicio).order_by(CatalogoServicio.orden, CatalogoServicio.nombre)
    ).all()
    return list(items)


@router.post("/admin/servicios-catalogo", response_model=CatalogoServicioRead, status_code=201)
def crear_servicio_catalogo(
    data: CatalogoServicioCreate,
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    nombre = data.nombre.strip()
    if not nombre:
        raise HTTPException(400, "El nombre es obligatorio")
    existente = session.exec(
        select(CatalogoServicio).where(func.lower(CatalogoServicio.nombre) == nombre.lower())
    ).first()
    if existente:
        raise HTTPException(400, "Ya existe un servicio con ese nombre")
    servicio = CatalogoServicio(nombre=nombre, cups=data.cups, activo=data.activo, orden=data.orden, permite_viaticos=data.permite_viaticos)
    session.add(servicio)
    session.commit()
    session.refresh(servicio)
    return servicio


@router.put("/admin/servicios-catalogo/{servicio_id}", response_model=CatalogoServicioRead)
def actualizar_servicio_catalogo(
    servicio_id: int,
    data: CatalogoServicioUpdate,
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    servicio = session.get(CatalogoServicio, servicio_id)
    if not servicio:
        raise HTTPException(404, "Servicio no encontrado")
    if data.nombre is not None:
        servicio.nombre = data.nombre.strip()
    if data.cups is not None:
        servicio.cups = data.cups.strip() or None
    if data.activo is not None:
        servicio.activo = data.activo
    if data.orden is not None:
        servicio.orden = data.orden
    if data.permite_viaticos is not None:
        servicio.permite_viaticos = data.permite_viaticos
    servicio.updated_at = datetime.utcnow()
    session.add(servicio)
    session.commit()
    session.refresh(servicio)
    return servicio


@router.delete("/admin/servicios-catalogo/{servicio_id}", status_code=204)
def eliminar_servicio_catalogo(
    servicio_id: int,
    session: Session = Depends(get_session),
    admin: Usuario = Depends(get_current_admin),
):
    servicio = session.get(CatalogoServicio, servicio_id)
    if not servicio:
        raise HTTPException(404, "Servicio no encontrado")
    session.delete(servicio)
    session.commit()
    return None
