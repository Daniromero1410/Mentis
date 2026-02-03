from openpyxl import load_workbook
from io import BytesIO
from typing import Optional
from datetime import date, datetime
import os
from pathlib import Path

# Ruta de la plantilla base
PLANTILLA_PATH = os.path.join(os.path.dirname(__file__), 'plantilla_valoracion.xlsx')


def sanitize_filename(text: str) -> str:
    """
    Limpia un texto para usarlo como nombre de archivo
    """
    # Reemplazar caracteres no permitidos
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        text = text.replace(char, '')
    # Reemplazar espacios con guiones bajos
    text = text.replace(' ', '_')
    # Limitar longitud
    if len(text) > 50:
        text = text[:50]
    return text


def safe_write_cell(ws, cell_ref, value):
    """
    Intenta escribir en una celda de forma segura.
    Si la celda es parte de una celda combinada, intenta escribir en la celda principal.
    Si falla, simplemente omite la escritura.
    """
    try:
        ws[cell_ref] = value
    except AttributeError as e:
        # Si es una MergedCell, intentar obtener la celda principal
        if 'MergedCell' in str(e):
            # Buscar la celda principal de la celda combinada
            for merged_range in ws.merged_cells.ranges:
                if cell_ref in merged_range:
                    # Escribir en la celda principal (top-left de la combinada)
                    ws.cell(merged_range.min_row, merged_range.min_col).value = value
                    return
        # Si no se puede escribir, simplemente continuar
        pass


def generar_excel_valoracion(
    valoracion,
    trabajador,
    info_laboral,
    historia_ocupacional: list,
    actividad_laboral,
    evaluaciones: list,
    concepto,
    guardar_archivo: bool = False,
    output_dir: str = "pdfs"
):
    """
    Genera un archivo Excel con la valoración completa usando la plantilla original

    Args:
        valoracion: Objeto valoración
        trabajador: Objeto trabajador
        info_laboral: Objeto info laboral
        historia_ocupacional: Lista de historia ocupacional
        actividad_laboral: Objeto actividad laboral
        evaluaciones: Lista de evaluaciones
        concepto: Objeto concepto
        guardar_archivo: Si True, guarda el archivo en disco y retorna la ruta. Si False, retorna BytesIO
        output_dir: Directorio donde guardar el archivo (solo si guardar_archivo=True)

    Returns:
        BytesIO con el contenido del Excel o ruta del archivo guardado
    """

    # Cargar la plantilla
    wb = load_workbook(PLANTILLA_PATH)
    ws = wb.active
    
    # ===== FECHA DE VALORACIÓN (Fila 7) =====
    if valoracion.fecha_valoracion:
        safe_write_cell(ws, 'R7', valoracion.fecha_valoracion.day)
        safe_write_cell(ws, 'T7', valoracion.fecha_valoracion.month)
        safe_write_cell(ws, 'V7', valoracion.fecha_valoracion.year)

    # ===== IDENTIFICACIÓN =====
    safe_write_cell(ws, 'H10', trabajador.nombre if trabajador else "")
    # CORRECCIÓN 1: Documento completo en H11 (merged H11:Z11)
    safe_write_cell(ws, 'H11', trabajador.documento if trabajador and trabajador.documento else "")
    

    # Fecha de nacimiento
    if trabajador and trabajador.fecha_nacimiento:
        safe_write_cell(ws, 'H13', trabajador.fecha_nacimiento.day)
        safe_write_cell(ws, 'J13', trabajador.fecha_nacimiento.month)
        safe_write_cell(ws, 'L13', trabajador.fecha_nacimiento.year)
        today = date.today()
        edad = today.year - trabajador.fecha_nacimiento.year - (
            (today.month, today.day) < (trabajador.fecha_nacimiento.month, trabajador.fecha_nacimiento.day)
        )
        safe_write_cell(ws, 'P13', edad)

    # Estado civil - Marcar con X
    if trabajador and trabajador.estado_civil:
        estado = trabajador.estado_civil.value if hasattr(trabajador.estado_civil, 'value') else str(trabajador.estado_civil)
        estado_map = {'casado': 'I15', 'soltero': 'K15', 'union_libre': 'N15', 'separado': 'S15', 'viudo': 'V15'}
        if estado in estado_map:
            safe_write_cell(ws, estado_map[estado], "X")

    # ===== NIVEL EDUCATIVO (Filas 16-18) =====
    if trabajador and trabajador.nivel_educativo:
        nivel_str = trabajador.nivel_educativo.value if hasattr(trabajador.nivel_educativo, 'value') else str(trabajador.nivel_educativo)
        nivel_lower = nivel_str.lower()

        # El nivel_educativo viene como string con valores separados por comas
        # Ejemplo: "Formación empírica, Básica primaria, Técnico/Tecnológico"
        # Revisar cada opción y marcar el checkbox correspondiente

        # Fila 16
        if 'formación empírica' in nivel_lower or 'formacion empirica' in nivel_lower:
            safe_write_cell(ws, 'L16', "X")
        if 'básica primaria' in nivel_lower or 'basica primaria' in nivel_lower:
            safe_write_cell(ws, 'R16', "X")
        if 'bachillerato vocacional' in nivel_lower or 'bachillerato: vocacional' in nivel_lower:
            safe_write_cell(ws, 'X16', "X")

        # Fila 17
        if 'bachillerato modalidad' in nivel_lower or 'bachillerato: modalidad' in nivel_lower:
            safe_write_cell(ws, 'L17', "X")
        if 'técnico' in nivel_lower or 'tecnológico' in nivel_lower:
            safe_write_cell(ws, 'R17', "X")
        if 'universitario' in nivel_lower:
            safe_write_cell(ws, 'X17', "X")

        # Fila 18
        if 'especialización' in nivel_lower or 'postgrado' in nivel_lower or 'maestría' in nivel_lower:
            safe_write_cell(ws, 'L18', "X")
        if 'formación informal' in nivel_lower or 'formacion informal' in nivel_lower:
            safe_write_cell(ws, 'R18', "X")
        if 'analfabeta' in nivel_lower:
            safe_write_cell(ws, 'X18', "X")

    # Formación específica (Fila 19)
    if trabajador and trabajador.formacion_especifica:
        safe_write_cell(ws, 'L19', trabajador.formacion_especifica)

    safe_write_cell(ws, 'H20', trabajador.telefonos if trabajador and trabajador.telefonos else "")
    safe_write_cell(ws, 'H21', trabajador.direccion if trabajador and trabajador.direccion else "")

    # Zona
    if trabajador and trabajador.zona:
        zona = trabajador.zona.value if hasattr(trabajador.zona, 'value') else str(trabajador.zona)
        if zona == 'urbano':
            safe_write_cell(ws, 'S21', "X")
        elif zona == 'rural':
            safe_write_cell(ws, 'U21', "X")

    safe_write_cell(ws, 'H22', trabajador.diagnostico_mental if trabajador and trabajador.diagnostico_mental else "")
    
    # ===== INFORMACIÓN LABORAL =====
    if info_laboral and info_laboral.fecha_evento_atel:
        safe_write_cell(ws, 'H26', info_laboral.fecha_evento_atel.strftime("%d/%m/%Y"))
    
    if info_laboral:
        if info_laboral.eventos_no_laborales:
            safe_write_cell(ws, 'H27', "X")
            if info_laboral.evento_no_laboral_fecha:
                safe_write_cell(ws, 'L27', info_laboral.evento_no_laboral_fecha.strftime("%d/%m/%Y"))
            safe_write_cell(ws, 'R27', info_laboral.evento_no_laboral_diagnostico or "")
        else:
            safe_write_cell(ws, 'J27', "X")
    
    safe_write_cell(ws, 'H28', info_laboral.eps if info_laboral and info_laboral.eps else "")
    safe_write_cell(ws, 'H29', info_laboral.fondo_pension if info_laboral and info_laboral.fondo_pension else "")
    safe_write_cell(ws, 'H30', info_laboral.dias_incapacidad if info_laboral and info_laboral.dias_incapacidad else "")
    safe_write_cell(ws, 'H31', info_laboral.empresa if info_laboral and info_laboral.empresa else "")
    
    if info_laboral:
        if info_laboral.vinculacion_laboral:
            safe_write_cell(ws, 'J32', "X")
        else:
            safe_write_cell(ws, 'H32', "X")
    
    safe_write_cell(ws, 'H33', info_laboral.tipo_vinculacion if info_laboral and info_laboral.tipo_vinculacion else "")
    
    if info_laboral and info_laboral.modalidad:
        modalidad = info_laboral.modalidad.value if hasattr(info_laboral.modalidad, 'value') else str(info_laboral.modalidad)
        if modalidad == 'presencial':
            safe_write_cell(ws, 'I34', "X")
        elif modalidad == 'teletrabajo':
            safe_write_cell(ws, 'N34', "X")
        elif modalidad == 'trabajo_en_casa':
            safe_write_cell(ws, 'T34', "X")
    
    safe_write_cell(ws, 'H35', info_laboral.tiempo_modalidad if info_laboral and info_laboral.tiempo_modalidad else "")
    safe_write_cell(ws, 'H36', info_laboral.nit_empresa if info_laboral and info_laboral.nit_empresa else "")
    
    if info_laboral and info_laboral.fecha_ingreso_empresa:
        safe_write_cell(ws, 'H37', info_laboral.fecha_ingreso_empresa.day)
        safe_write_cell(ws, 'J37', info_laboral.fecha_ingreso_empresa.month)
        safe_write_cell(ws, 'L37', info_laboral.fecha_ingreso_empresa.year)
    
    if info_laboral:
        safe_write_cell(ws, 'P38', info_laboral.antiguedad_empresa_anos or "")
        safe_write_cell(ws, 'V38', info_laboral.antiguedad_empresa_meses or "")
        safe_write_cell(ws, 'P42', info_laboral.antiguedad_cargo_anos or "")
        safe_write_cell(ws, 'V42', info_laboral.antiguedad_cargo_meses or "")
    
    contacto = ""
    if info_laboral:
        if info_laboral.contacto_empresa:
            contacto = info_laboral.contacto_empresa
        if info_laboral.cargo_contacto:
            contacto += f" / {info_laboral.cargo_contacto}" if contacto else info_laboral.cargo_contacto
    safe_write_cell(ws, 'H43',contacto)
    safe_write_cell(ws, 'H44', info_laboral.correos if info_laboral and info_laboral.correos else "")
    safe_write_cell(ws, 'H45', info_laboral.telefonos_empresa if info_laboral and info_laboral.telefonos_empresa else "")
    
    # ===== HISTORIA OCUPACIONAL (Filas 49-51) =====
    for i, hist in enumerate(historia_ocupacional[:3]):
        fila = 49 + i
        safe_write_cell(ws, f'A{fila}', hist.empresa or "")
        safe_write_cell(ws, f'F{fila}', hist.cargo_funciones or "")
        safe_write_cell(ws, f'L{fila}', hist.duracion or "")
        safe_write_cell(ws, f'R{fila}', hist.motivo_retiro or "")
    
    # CORRECCIÓN 3: Otros oficios en A52 y A53 (merged A:Z)
    safe_write_cell(ws, 'A52', actividad_laboral.otros_oficios if actividad_laboral and actividad_laboral.otros_oficios else "")
    safe_write_cell(ws, 'A53', actividad_laboral.oficios_interes if actividad_laboral and actividad_laboral.oficios_interes else "")
    
    # ===== ACTIVIDAD LABORAL ACTUAL =====
    # CORRECCIÓN 4: Actividad laboral actual en A55-A59 (merged A:Z)
    safe_write_cell(ws, 'A55', actividad_laboral.nombre_cargo if actividad_laboral and actividad_laboral.nombre_cargo else "")
    safe_write_cell(ws, 'A56', actividad_laboral.tareas if actividad_laboral and actividad_laboral.tareas else "")
    safe_write_cell(ws, 'A57', actividad_laboral.herramientas if actividad_laboral and actividad_laboral.herramientas else "")
    safe_write_cell(ws, 'A58', actividad_laboral.horario if actividad_laboral and actividad_laboral.horario else "")
    safe_write_cell(ws, 'A59', actividad_laboral.elementos_proteccion if actividad_laboral and actividad_laboral.elementos_proteccion else "")
    
    # ===== FACTORES DE RIESGO PSICOSOCIALES =====
    items_filas = {
        ('demandas_cuantitativas', 1): 62, ('demandas_cuantitativas', 2): 63,
        ('demandas_cuantitativas', 3): 64, ('demandas_cuantitativas', 4): 65,
        ('demandas_carga_mental', 1): 67, ('demandas_carga_mental', 2): 68,
        ('demandas_carga_mental', 3): 69, ('demandas_carga_mental', 4): 70,
        ('demandas_carga_mental', 5): 71, ('demandas_carga_mental', 6): 72,
        ('demandas_carga_mental', 7): 73, ('demandas_carga_mental', 8): 74,
        ('demandas_emocionales', 1): 76, ('demandas_emocionales', 2): 77,
        ('demandas_emocionales', 3): 78, ('demandas_emocionales', 4): 79,
        ('demandas_emocionales', 5): 80, ('demandas_emocionales', 6): 81,
        ('exigencias_responsabilidad', 1): 83, ('exigencias_responsabilidad', 2): 84,
        ('exigencias_responsabilidad', 3): 85, ('exigencias_responsabilidad', 4): 86,
        ('exigencias_responsabilidad', 5): 87, ('exigencias_responsabilidad', 6): 88,
        ('consistencia_rol', 1): 90, ('consistencia_rol', 2): 91,
        ('consistencia_rol', 3): 92, ('consistencia_rol', 4): 93,
        ('consistencia_rol', 5): 94, ('consistencia_rol', 6): 95,
        ('consistencia_rol', 7): 96,
        ('demandas_ambientales', 1): 98, ('demandas_ambientales', 2): 99,
        ('demandas_ambientales', 3): 100, ('demandas_ambientales', 4): 101,
        ('demandas_ambientales', 5): 102, ('demandas_ambientales', 6): 103,
        ('demandas_ambientales', 7): 104, ('demandas_ambientales', 8): 105,
        ('demandas_ambientales', 9): 106, ('demandas_ambientales', 10): 107,
        ('demandas_jornada', 1): 109, ('demandas_jornada', 2): 110,
    }
    
    for ev in evaluaciones:
        cat = ev.categoria.value if hasattr(ev.categoria, 'value') else str(ev.categoria)
        key = (cat, ev.item_numero)
        
        if key in items_filas:
            fila = items_filas[key]
            if ev.calificacion:
                calif = ev.calificacion.value if hasattr(ev.calificacion, 'value') else str(ev.calificacion)
                if calif == 'alto':
                    safe_write_cell(ws, f'L{fila}', "X")
                elif calif == 'medio':
                    safe_write_cell(ws, f'N{fila}', "X")
                elif calif == 'bajo':
                    safe_write_cell(ws, f'P{fila}', "X")
            if ev.observaciones:
                safe_write_cell(ws, f'T{fila}', ev.observaciones)
    
    # ===== CONCEPTO PSICOLÓGICO FINAL =====
    if concepto:
        concepto_texto = concepto.concepto_editado or concepto.concepto_generado or ""
        safe_write_cell(ws, 'A112', concepto_texto)
    
    if concepto and concepto.orientacion_reintegro:
        safe_write_cell(ws, 'A114', concepto.orientacion_reintegro)
    
    # ===== FIRMAS =====
    if concepto:
        safe_write_cell(ws, 'A117', concepto.elaboro_nombre or "")
        safe_write_cell(ws, 'J117', concepto.reviso_nombre or "")

    # Guardar según el modo
    if guardar_archivo:
        # Guardar como archivo en disco con nombre personalizado
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)

        # Generar nombre de archivo con nombre y documento del trabajador
        nombre_sanitizado = sanitize_filename(trabajador.nombre if trabajador else "sin_nombre")
        documento_sanitizado = sanitize_filename(trabajador.documento if trabajador else "sin_doc")
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"valoracion_{nombre_sanitizado}_{documento_sanitizado}_{timestamp}.xlsx"
        filepath = output_path / filename

        wb.save(str(filepath))
        return str(filepath)
    else:
        # Retornar como BytesIO para descarga directa
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


def generar_plantilla_vacia() -> BytesIO:
    """Genera una copia de la plantilla vacía original"""
    wb = load_workbook(PLANTILLA_PATH)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
