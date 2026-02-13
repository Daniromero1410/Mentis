"""
Generador de Excel simplificado para valoraciones ocupacionales
Usa plantilla_valoracion_simple.xlsx que es mucho más fácil de mapear
"""
from openpyxl import load_workbook
from io import BytesIO
from typing import Optional, List
from datetime import date, datetime
import os
from pathlib import Path

# Ruta de la plantilla simplificada (la original compleja está en plantilla_valoracion.xlsx)
PLANTILLA_PATH = os.path.join(os.path.dirname(__file__), 'plantilla_valoracion_simple.xlsx')


def sanitize_filename(text: str) -> str:
    """Limpia un texto para usarlo como nombre de archivo"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        text = text.replace(char, '')
    text = text.replace(' ', '_')
    if len(text) > 50:
        text = text[:50]
    return text


def safe_write_cell(ws, cell_ref, value):
    """Escribe en una celda de forma segura"""
    try:
        ws[cell_ref] = value
    except Exception:
        pass


def generar_excel_valoracion(
    valoracion,
    trabajador,
    info_laboral,
    historia_ocupacional: List,
    actividad_laboral,
    evaluaciones: List,
    concepto,
    guardar_archivo: bool = True,
    output_dir: str = "app/services"
) -> Optional[BytesIO]:
    """
    Genera un archivo Excel de valoración ocupacional usando plantilla simplificada

    Returns:
        - Si guardar_archivo=True: ruta del archivo guardado
        - Si guardar_archivo=False: BytesIO con el contenido
    """

    # Cargar plantilla
    wb = load_workbook(PLANTILLA_PATH)
    ws = wb.active

    # ===== FECHA DE VALORACIÓN (Fila 2) =====
    if valoracion and valoracion.fecha_valoracion:
        safe_write_cell(ws, 'B2', valoracion.fecha_valoracion.day)
        safe_write_cell(ws, 'C2', valoracion.fecha_valoracion.month)
        safe_write_cell(ws, 'D2', valoracion.fecha_valoracion.year)

    # ===== SECCIÓN 1: IDENTIFICACIÓN (Inicia en fila 5) =====
    # Nombre (fila 5)
    safe_write_cell(ws, 'B5', trabajador.nombre if trabajador else "")

    # Documento (fila 6)
    safe_write_cell(ws, 'B6', trabajador.documento if trabajador else "")

    # Fecha de nacimiento y edad (fila 7)
    if trabajador and trabajador.fecha_nacimiento:
        safe_write_cell(ws, 'B7', trabajador.fecha_nacimiento.day)
        safe_write_cell(ws, 'C7', trabajador.fecha_nacimiento.month)
        safe_write_cell(ws, 'D7', trabajador.fecha_nacimiento.year)

        # Calcular edad
        today = date.today()
        edad = today.year - trabajador.fecha_nacimiento.year - (
            (today.month, today.day) < (trabajador.fecha_nacimiento.month, trabajador.fecha_nacimiento.day)
        )
        safe_write_cell(ws, 'G7', edad)

    # Sexo (fila 8) - Marcar con X
    if trabajador and trabajador.sexo:
        sexo = trabajador.sexo.value if hasattr(trabajador.sexo, 'value') else str(trabajador.sexo)
        if sexo.lower() in ['m', 'masculino']:
            safe_write_cell(ws, 'B8', 'M [X]')
            safe_write_cell(ws, 'C8', 'F [ ]')
        else:
            safe_write_cell(ws, 'B8', 'M [ ]')
            safe_write_cell(ws, 'C8', 'F [X]')

    # Estado civil (fila 9) - Marcar con X
    if trabajador and trabajador.estado_civil:
        estado = trabajador.estado_civil.value if hasattr(trabajador.estado_civil, 'value') else str(trabajador.estado_civil)
        estados = {
            'casado': 'B9',
            'soltero': 'C9',
            'union_libre': 'D9',
            'separado': 'E9',
            'viudo': 'F9'
        }
        # Limpiar todas las marcas primero
        for col in ['B9', 'C9', 'D9', 'E9', 'F9']:
            val = ws[col].value
            if val:
                safe_write_cell(ws, col, val.replace('[X]', '[ ]'))

        # Marcar el correspondiente
        if estado in estados:
            val = ws[estados[estado]].value
            if val:
                safe_write_cell(ws, estados[estado], val.replace('[ ]', '[X]'))

    # Nivel educativo (fila 10) - Texto completo separado por comas
    if trabajador and trabajador.nivel_educativo:
        nivel_str = trabajador.nivel_educativo.value if hasattr(trabajador.nivel_educativo, 'value') else str(trabajador.nivel_educativo)
        safe_write_cell(ws, 'B10', nivel_str)

    # Formación específica (fila 11)
    if trabajador and trabajador.formacion_especifica:
        safe_write_cell(ws, 'B11', trabajador.formacion_especifica)

    # Teléfonos (fila 12)
    safe_write_cell(ws, 'B12', trabajador.telefonos if trabajador and trabajador.telefonos else "")

    # Dirección (fila 13)
    safe_write_cell(ws, 'B13', trabajador.direccion if trabajador and trabajador.direccion else "")

    # Zona (fila 14) - Marcar con X
    if trabajador and trabajador.zona:
        zona = trabajador.zona.value if hasattr(trabajador.zona, 'value') else str(trabajador.zona)
        if zona.lower() == 'urbana':
            safe_write_cell(ws, 'B14', 'Urbana [X]')
            safe_write_cell(ws, 'C14', 'Rural [ ]')
        else:
            safe_write_cell(ws, 'B14', 'Urbana [ ]')
            safe_write_cell(ws, 'C14', 'Rural [X]')

    # Ciudad, Departamento, País (fila 15)
    if trabajador:
        safe_write_cell(ws, 'B15', trabajador.ciudad if trabajador.ciudad else "")
        safe_write_cell(ws, 'D15', trabajador.departamento if trabajador.departamento else "")
        safe_write_cell(ws, 'F15', trabajador.pais if trabajador.pais else "")

    # ===== SECCIÓN 2: INFO LABORAL (Inicia en fila 18) =====
    if info_laboral:
        # Empresa (fila 18)
        safe_write_cell(ws, 'B18', info_laboral.empresa if info_laboral.empresa else "")

        # NIT (fila 19)
        safe_write_cell(ws, 'B19', info_laboral.nit if info_laboral.nit else "")

        # Ciudad empresa y Departamento (fila 20)
        safe_write_cell(ws, 'B20', info_laboral.ciudad if info_laboral.ciudad else "")
        safe_write_cell(ws, 'D20', info_laboral.departamento if info_laboral.departamento else "")

        # Área (fila 21)
        safe_write_cell(ws, 'B21', info_laboral.area if info_laboral.area else "")

        # Cargo (fila 22)
        safe_write_cell(ws, 'B22', info_laboral.cargo if info_laboral.cargo else "")

        # Antigüedad (fila 23)
        safe_write_cell(ws, 'B23', info_laboral.antiguedad_cargo if info_laboral.antiguedad_cargo else "")
        safe_write_cell(ws, 'D23', info_laboral.antiguedad_empresa if info_laboral.antiguedad_empresa else "")

        # Tipo de contrato (fila 24)
        safe_write_cell(ws, 'B24', info_laboral.tipo_contrato if info_laboral.tipo_contrato else "")

    # ===== SECCIÓN 3: HISTORIA OCUPACIONAL (Inicia en fila 26) =====
    # Encabezado en fila 27, Filas de datos: 28, 29, 30 (3 empleos anteriores)
    for i, hist in enumerate(historia_ocupacional[:3]):
        fila = 28 + i
        safe_write_cell(ws, f'A{fila}', hist.empresa if hist.empresa else "")
        safe_write_cell(ws, f'B{fila}', hist.cargo_funciones if hist.cargo_funciones else "")
        safe_write_cell(ws, f'E{fila}', hist.duracion if hist.duracion else "")
        safe_write_cell(ws, f'G{fila}', hist.motivo_retiro if hist.motivo_retiro else "")

    # Otros oficios (fila 32)
    if actividad_laboral and actividad_laboral.otros_oficios:
        safe_write_cell(ws, 'B32', actividad_laboral.otros_oficios)

    # Oficios de interés (fila 33)
    if actividad_laboral and actividad_laboral.oficios_interes:
        safe_write_cell(ws, 'B33', actividad_laboral.oficios_interes)

    # ===== SECCIÓN 4: ACTIVIDAD LABORAL ACTUAL (Inicia en fila 35) =====
    if actividad_laboral:
        # Nombre del cargo (fila 36)
        safe_write_cell(ws, 'B36', actividad_laboral.nombre_cargo if actividad_laboral.nombre_cargo else "")

        # Tareas (fila 37)
        safe_write_cell(ws, 'B37', actividad_laboral.tareas if actividad_laboral.tareas else "")

        # Herramientas (fila 38)
        safe_write_cell(ws, 'B38', actividad_laboral.herramientas if actividad_laboral.herramientas else "")

        # Horario (fila 39)
        safe_write_cell(ws, 'B39', actividad_laboral.horario if actividad_laboral.horario else "")

        # EPP (fila 40)
        safe_write_cell(ws, 'B40', actividad_laboral.elementos_proteccion if actividad_laboral.elementos_proteccion else "")

    # ===== SECCIÓN 5: FACTORES DE RIESGO (Inicia en fila 42) =====
    # Mapeo de categorías a filas (categoría en una fila, items empiezan en la siguiente)
    categorias_filas = {
        'demandas_cuantitativas': (45, 4),  # (fila_inicio_items, num_items)
        'demandas_carga_mental': (50, 8),
        'demandas_emocionales': (59, 6),
        'exigencias_responsabilidad': (66, 6),
        'consistencia_rol': (73, 7),
        'demandas_ambientales': (81, 14),
        'demandas_jornada': (96, 4),
    }

    # Llenar evaluaciones
    for eval in evaluaciones:
        categoria = eval.categoria
        item_numero = eval.item_numero
        respuesta = eval.respuesta

        if categoria in categorias_filas:
            fila_inicio, num_items = categorias_filas[categoria]

            # Calcular fila del item
            fila = fila_inicio + (item_numero - 1)

            # Limpiar primero
            safe_write_cell(ws, f'G{fila}', '')
            safe_write_cell(ws, f'H{fila}', '')
            safe_write_cell(ws, f'I{fila}', '')

            # Marcar respuesta
            if respuesta == 'si':
                safe_write_cell(ws, f'G{fila}', 'X')
            elif respuesta == 'no':
                safe_write_cell(ws, f'H{fila}', 'X')
            elif respuesta == 'na':
                safe_write_cell(ws, f'I{fila}', 'X')

    # ===== SECCIÓN 6: CONCEPTO FINAL (Inicia en fila 101) =====
    if concepto:
        # Apariencia personal (fila 102)
        safe_write_cell(ws, 'B102', concepto.apariencia_personal if concepto.apariencia_personal else "")

        # Actitud (fila 103)
        safe_write_cell(ws, 'B103', concepto.actitud if concepto.actitud else "")

        # Concepto (fila 104)
        safe_write_cell(ws, 'B104', concepto.concepto if concepto.concepto else "")

        # Recomendaciones (fila 105)
        safe_write_cell(ws, 'B105', concepto.recomendaciones if concepto.recomendaciones else "")

    # Nombre del evaluador (fila 107)
    if valoracion and valoracion.evaluador_nombre:
        safe_write_cell(ws, 'B107', valoracion.evaluador_nombre)

    # Fecha de firma (fila 109)
    if valoracion and valoracion.fecha_valoracion:
        safe_write_cell(ws, 'B109', valoracion.fecha_valoracion.day)
        safe_write_cell(ws, 'C109', valoracion.fecha_valoracion.month)
        safe_write_cell(ws, 'D109', valoracion.fecha_valoracion.year)

    # ===== GUARDAR O RETORNAR =====
    if guardar_archivo:
        # Generar nombre de archivo
        nombre = trabajador.nombre if trabajador and trabajador.nombre else "sin_nombre"
        documento = trabajador.documento if trabajador and trabajador.documento else "sin_documento"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        filename = f"valoracion_{sanitize_filename(nombre)}_{sanitize_filename(documento)}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Crear directorio si no existe
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        # Guardar
        wb.save(filepath)
        return filepath
    else:
        # Retornar BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


def generar_plantilla_vacia() -> BytesIO:
    """Genera una copia de la plantilla vacía simplificada"""
    wb = load_workbook(PLANTILLA_PATH)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
