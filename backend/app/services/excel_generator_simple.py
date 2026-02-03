"""
Generador de Excel simple para valoraciones
Crea un Excel básico listando todos los datos del formulario
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
from pathlib import Path
import os


def calcular_edad(fecha_nacimiento: date) -> int:
    """Calcula la edad a partir de la fecha de nacimiento"""
    today = date.today()
    edad = today.year - fecha_nacimiento.year - (
        (today.month, today.day) < (fecha_nacimiento.month, fecha_nacimiento.day)
    )
    return edad


def sanitize_filename(text: str) -> str:
    """Limpia un texto para usarlo como nombre de archivo"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        text = text.replace(char, '')
    text = text.replace(' ', '_')
    if len(text) > 50:
        text = text[:50]
    return text


def generar_excel_valoracion(
    valoracion,
    trabajador,
    info_laboral,
    historia_ocupacional,
    actividad_laboral,
    evaluaciones,
    concepto,
    guardar_archivo: bool = True,
    output_dir: str = "pdfs"
) -> str:
    """
    Genera un Excel simple con todos los datos de la valoración

    Returns:
        Ruta del archivo Excel generado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Valoración"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=11)
    label_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # Título
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "EVALUACIÓN PSICOLÓGICA - REINTEGRO Y REUBICACIÓN LABORAL"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    row += 2

    # DATOS GENERALES
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "DATOS GENERALES"
    cell.font = section_font
    cell.fill = section_fill
    row += 1

    # Fecha valoración
    ws[f'A{row}'] = "Fecha de valoración:"
    ws[f'A{row}'].font = label_font
    ws[f'B{row}'] = valoracion.fecha_valoracion.strftime('%d/%m/%Y') if valoracion and valoracion.fecha_valoracion else ''
    row += 1

    if trabajador:
        # Nombre
        ws[f'A{row}'] = "Nombres y apellidos:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = trabajador.nombre or ''
        row += 1

        # Documento
        ws[f'A{row}'] = "Documento:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = trabajador.documento or ''
        row += 1

        # ID Siniestro
        ws[f'A{row}'] = "Identificación siniestro:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = trabajador.identificacion_siniestro or ''
        row += 1

        # Fecha nacimiento y edad
        if trabajador.fecha_nacimiento:
            ws[f'A{row}'] = "Fecha de nacimiento:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = trabajador.fecha_nacimiento.strftime('%d/%m/%Y')
            row += 1

            ws[f'A{row}'] = "Edad:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = calcular_edad(trabajador.fecha_nacimiento)
            row += 1

        # Estado civil
        if trabajador.estado_civil:
            ws[f'A{row}'] = "Estado civil:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = trabajador.estado_civil.value.replace('_', ' ').title()
            row += 1

        # Nivel educativo
        ws[f'A{row}'] = "Nivel educativo:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = trabajador.nivel_educativo or ''
        row += 1

        # Formación específica
        ws[f'A{row}'] = "Formación específica:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = trabajador.formacion_especifica or ''
        row += 1

        # Teléfonos
        ws[f'A{row}'] = "Teléfonos:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = trabajador.telefonos or ''
        row += 1

        # Dirección
        ws[f'A{row}'] = "Dirección:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = trabajador.direccion or ''
        row += 1

        # Zona
        if trabajador.zona:
            ws[f'A{row}'] = "Zona:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = trabajador.zona.value.title()
            row += 1

        # Diagnóstico mental
        ws[f'A{row}'] = "Diagnóstico mental:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = trabajador.diagnostico_mental or ''
        row += 1

    row += 1

    # INFORMACIÓN LABORAL
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = "INFORMACIÓN LABORAL"
    cell.font = section_font
    cell.fill = section_fill
    row += 1

    if info_laboral:
        # Empresa
        ws[f'A{row}'] = "Empresa:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = info_laboral.empresa or ''
        row += 1

        # NIT
        ws[f'A{row}'] = "NIT:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = info_laboral.nit_empresa or ''
        row += 1

        # EPS
        ws[f'A{row}'] = "EPS:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = info_laboral.eps or ''
        row += 1

        # Fondo pensión
        ws[f'A{row}'] = "Fondo de pensión:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = info_laboral.fondo_pension or ''
        row += 1

        # Tipo vinculación
        ws[f'A{row}'] = "Tipo de vinculación:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = info_laboral.tipo_vinculacion or ''
        row += 1

        # Modalidad
        if info_laboral.modalidad:
            ws[f'A{row}'] = "Modalidad:"
            ws[f'A{row}'].font = label_font
            ws[f'B{row}'] = info_laboral.modalidad.value.replace('_', ' ').title()
            row += 1

        # Antigüedad empresa
        ws[f'A{row}'] = "Antigüedad en la empresa:"
        ws[f'A{row}'].font = label_font
        anos = info_laboral.antiguedad_empresa_anos or 0
        meses = info_laboral.antiguedad_empresa_meses or 0
        ws[f'B{row}'] = f"{anos} años, {meses} meses"
        row += 1

        # Antigüedad cargo
        ws[f'A{row}'] = "Antigüedad en el cargo:"
        ws[f'A{row}'].font = label_font
        anos = info_laboral.antiguedad_cargo_anos or 0
        meses = info_laboral.antiguedad_cargo_meses or 0
        ws[f'B{row}'] = f"{anos} años, {meses} meses"
        row += 1

    row += 1

    # HISTORIA OCUPACIONAL
    if historia_ocupacional and len(historia_ocupacional) > 0:
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "HISTORIA OCUPACIONAL"
        cell.font = section_font
        cell.fill = section_fill
        row += 1

        # Encabezados
        ws[f'A{row}'] = "Empresa"
        ws[f'B{row}'] = "Cargo/Funciones"
        ws[f'C{row}'] = "Duración"
        ws[f'D{row}'] = "Motivo de retiro"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = label_font
        row += 1

        # Datos
        for hist in historia_ocupacional:
            ws[f'A{row}'] = hist.empresa or ''
            ws[f'B{row}'] = hist.cargo_funciones or ''
            ws[f'C{row}'] = hist.duracion or ''
            ws[f'D{row}'] = hist.motivo_retiro or ''
            row += 1

        row += 1

    # ACTIVIDAD LABORAL ACTUAL
    if actividad_laboral:
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "ACTIVIDAD LABORAL ACTUAL"
        cell.font = section_font
        cell.fill = section_fill
        row += 1

        ws[f'A{row}'] = "Cargo:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = actividad_laboral.nombre_cargo or ''
        row += 1

        ws[f'A{row}'] = "Tareas:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = actividad_laboral.tareas or ''
        row += 1

        ws[f'A{row}'] = "Herramientas:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = actividad_laboral.herramientas or ''
        row += 1

        ws[f'A{row}'] = "Horario:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = actividad_laboral.horario or ''
        row += 1

        row += 1

    # FACTORES DE RIESGO PSICOSOCIAL
    if evaluaciones and len(evaluaciones) > 0:
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "FACTORES DE RIESGO PSICOSOCIAL"
        cell.font = section_font
        cell.fill = section_fill
        row += 1

        # Encabezados
        ws[f'A{row}'] = "Categoría"
        ws[f'B{row}'] = "Item"
        ws[f'C{row}'] = "Calificación"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = label_font
        row += 1

        # Datos
        for eval in evaluaciones:
            categoria = eval.categoria.value.replace('_', ' ').title() if eval.categoria else ''
            ws[f'A{row}'] = categoria
            ws[f'B{row}'] = eval.item_texto or ''
            ws[f'C{row}'] = eval.calificacion.value.upper() if eval.calificacion else ''
            row += 1

        row += 1

    # CONCEPTO FINAL
    if concepto:
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "CONCEPTO Y RECOMENDACIONES"
        cell.font = section_font
        cell.fill = section_fill
        row += 1

        ws[f'A{row}'] = "Concepto:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = concepto.concepto_editado or concepto.concepto_generado or ''
        row += 1

        ws[f'A{row}'] = "Orientación para reintegro:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = concepto.orientacion_reintegro or ''
        row += 1

        ws[f'A{row}'] = "Elaboró:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = concepto.elaboro_nombre or ''
        row += 1

        ws[f'A{row}'] = "Revisó:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = concepto.reviso_nombre or ''
        row += 1

    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30

    # Guardar
    if guardar_archivo:
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True, parents=True)

        nombre = trabajador.nombre if trabajador else "sin_nombre"
        documento = trabajador.documento if trabajador else "sin_documento"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        filename = f"valoracion_{sanitize_filename(nombre)}_{sanitize_filename(documento)}_{timestamp}.xlsx"
        filepath = output_path / filename

        wb.save(str(filepath))
        return str(filepath)

    return ""
