"""
Generador de Excel para valoraciones ocupacionales
Convierte los datos del formulario web en un documento Excel estructurado y profesional
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Optional, List
from datetime import date, datetime
import os
from pathlib import Path


def sanitize_filename(text: str) -> str:
    """Limpia un texto para usarlo como nombre de archivo"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        text = text.replace(char, '')
    text = text.replace(' ', '_')
    if len(text) > 50:
        text = text[:50]
    return text


def generar_excel_valoracion(
    valoracion,
    trabajador,
    info_laboral,
    historia_ocupacional: List,
    actividad_laboral,
    evaluaciones: List,
    concepto,
    guardar_archivo: bool = True,
    output_dir: str = "app/services"
) -> Optional[BytesIO]:
    """
    Genera un archivo Excel de valoración ocupacional desde cero

    Returns:
        - Si guardar_archivo=True: ruta del archivo guardado
        - Si guardar_archivo=False: BytesIO con el contenido
    """

    # Crear nuevo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Valoración Ocupacional"

    # Estilos
    titulo_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    titulo_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')

    seccion_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    seccion_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    label_font = Font(name='Arial', size=10, bold=True)
    label_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    data_font = Font(name='Arial', size=10)

    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Configurar anchos de columnas
    ws.column_dimensions['A'].width = 30
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 20

    row = 1

    # ===== ENCABEZADO =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = 'VALORACIÓN OCUPACIONAL'
    cell.font = titulo_font
    cell.fill = titulo_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 30
    row += 1

    # Fecha de valoración
    if valoracion and valoracion.fecha_valoracion:
        ws[f'A{row}'] = 'Fecha de valoración:'
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = valoracion.fecha_valoracion.strftime('%d/%m/%Y')
        ws[f'B{row}'].font = data_font
    row += 2

    # ===== SECCIÓN 1: DATOS DEL TRABAJADOR =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '1. DATOS DEL TRABAJADOR'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    if trabajador:
        campos_trabajador = [
            ('Nombre completo:', trabajador.nombre),
            ('Documento:', trabajador.documento),
            ('Identificación siniestro:', trabajador.identificacion_siniestro),
            ('Fecha de nacimiento:', trabajador.fecha_nacimiento.strftime('%d/%m/%Y') if trabajador.fecha_nacimiento else ''),
            ('Edad:', calcular_edad(trabajador.fecha_nacimiento) if trabajador.fecha_nacimiento else ''),
            ('Estado civil:', trabajador.estado_civil.value if trabajador.estado_civil else ''),
            ('Nivel educativo:', trabajador.nivel_educativo),
            ('Formación específica:', trabajador.formacion_especifica),
            ('Teléfonos:', trabajador.telefonos),
            ('Dirección:', trabajador.direccion),
            ('Zona:', trabajador.zona.value if trabajador.zona else ''),
            ('Diagnóstico mental:', trabajador.diagnostico_mental),
        ]

        for label, valor in campos_trabajador:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].fill = label_fill
            ws.merge_cells(f'B{row}:I{row}')
            ws[f'B{row}'] = valor or ''
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = border_thin
            row += 1

    row += 1

    # ===== SECCIÓN 2: INFORMACIÓN LABORAL =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '2. INFORMACIÓN LABORAL'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    if info_laboral:
        campos_info_laboral = [
            ('Empresa:', info_laboral.empresa),
            ('NIT empresa:', info_laboral.nit_empresa),
            ('Fecha evento AT/EL:', info_laboral.fecha_evento_atel.strftime('%d/%m/%Y') if info_laboral.fecha_evento_atel else ''),
            ('Eventos no laborales:', 'Sí' if info_laboral.eventos_no_laborales else 'No'),
            ('Fecha evento no laboral:', info_laboral.evento_no_laboral_fecha.strftime('%d/%m/%Y') if info_laboral.evento_no_laboral_fecha else ''),
            ('Diagnóstico evento no laboral:', info_laboral.evento_no_laboral_diagnostico),
            ('EPS:', info_laboral.eps),
            ('Fondo de pensión:', info_laboral.fondo_pension),
            ('Días de incapacidad:', info_laboral.dias_incapacidad),
            ('Vinculación laboral:', 'Sí' if info_laboral.vinculacion_laboral else 'No'),
            ('Tipo de vinculación:', info_laboral.tipo_vinculacion),
            ('Modalidad de trabajo:', info_laboral.modalidad.value if info_laboral.modalidad else ''),
            ('Tiempo en modalidad:', info_laboral.tiempo_modalidad),
            ('Fecha ingreso empresa:', info_laboral.fecha_ingreso_empresa.strftime('%d/%m/%Y') if info_laboral.fecha_ingreso_empresa else ''),
            ('Antigüedad en empresa:', f'{info_laboral.antiguedad_empresa_anos or 0} años, {info_laboral.antiguedad_empresa_meses or 0} meses'),
            ('Antigüedad en cargo:', f'{info_laboral.antiguedad_cargo_anos or 0} años, {info_laboral.antiguedad_cargo_meses or 0} meses'),
            ('Contacto empresa:', info_laboral.contacto_empresa),
            ('Cargo del contacto:', info_laboral.cargo_contacto),
            ('Correos:', info_laboral.correos),
            ('Teléfonos empresa:', info_laboral.telefonos_empresa),
        ]

        for label, valor in campos_info_laboral:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].fill = label_fill
            ws.merge_cells(f'B{row}:I{row}')
            ws[f'B{row}'] = str(valor) if valor is not None else ''
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = border_thin
            row += 1

    row += 1

    # ===== SECCIÓN 3: HISTORIA OCUPACIONAL =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '3. HISTORIA OCUPACIONAL'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    if historia_ocupacional:
        # Encabezados de tabla
        headers = ['Empresa', 'Cargo/Funciones', 'Duración', 'Motivo de retiro']
        cols = ['A', 'C', 'F', 'H']
        for i, header in enumerate(headers):
            ws[f'{cols[i]}{row}'] = header
            ws[f'{cols[i]}{row}'].font = label_font
            ws[f'{cols[i]}{row}'].fill = label_fill
            ws[f'{cols[i]}{row}'].border = border_thin
            ws[f'{cols[i]}{row}'].alignment = Alignment(horizontal='center')

        # Merged cells para cada columna
        ws.merge_cells(f'A{row}:B{row}')  # Empresa
        ws.merge_cells(f'C{row}:E{row}')  # Cargo
        ws.merge_cells(f'F{row}:G{row}')  # Duración
        ws.merge_cells(f'H{row}:I{row}')  # Motivo
        row += 1

        # Datos
        for hist in historia_ocupacional[:5]:  # Máximo 5 empleos
            ws.merge_cells(f'A{row}:B{row}')
            ws[f'A{row}'] = hist.empresa or ''
            ws[f'A{row}'].border = border_thin

            ws.merge_cells(f'C{row}:E{row}')
            ws[f'C{row}'] = hist.cargo_funciones or ''
            ws[f'C{row}'].border = border_thin

            ws.merge_cells(f'F{row}:G{row}')
            ws[f'F{row}'] = hist.duracion or ''
            ws[f'F{row}'].border = border_thin

            ws.merge_cells(f'H{row}:I{row}')
            ws[f'H{row}'] = hist.motivo_retiro or ''
            ws[f'H{row}'].border = border_thin

            row += 1

    row += 1

    # ===== SECCIÓN 4: ACTIVIDAD LABORAL ACTUAL =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '4. ACTIVIDAD LABORAL ACTUAL'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    if actividad_laboral:
        campos_actividad = [
            ('Nombre del cargo:', actividad_laboral.nombre_cargo),
            ('Tareas que realiza:', actividad_laboral.tareas),
            ('Herramientas/equipos:', actividad_laboral.herramientas),
            ('Horario de trabajo:', actividad_laboral.horario),
            ('Elementos de protección:', actividad_laboral.elementos_proteccion),
            ('Otros oficios u ocupaciones:', actividad_laboral.otros_oficios),
            ('Oficios de interés:', actividad_laboral.oficios_interes),
        ]

        for label, valor in campos_actividad:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].fill = label_fill
            ws.merge_cells(f'B{row}:I{row}')
            ws[f'B{row}'] = valor or ''
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = border_thin
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 30
            row += 1

    row += 1

    # ===== SECCIÓN 5: FACTORES DE RIESGO PSICOSOCIALES =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '5. FACTORES DE RIESGO PSICOSOCIALES'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    if evaluaciones:
        # Encabezados
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'Factor de Riesgo'
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].fill = label_fill
        ws[f'A{row}'].border = border_thin
        ws[f'A{row}'].alignment = Alignment(horizontal='center')

        for col, val in [('G', 'SI'), ('H', 'NO'), ('I', 'N/A')]:
            ws[f'{col}{row}'] = val
            ws[f'{col}{row}'].font = label_font
            ws[f'{col}{row}'].fill = label_fill
            ws[f'{col}{row}'].border = border_thin
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        row += 1

        # Agrupar por categoría
        categorias_dict = {}
        for eval in evaluaciones:
            if eval.categoria not in categorias_dict:
                categorias_dict[eval.categoria] = []
            categorias_dict[eval.categoria].append(eval)

        # Nombres de categorías
        nombres_categorias = {
            'demandas_cuantitativas': 'DEMANDAS CUANTITATIVAS',
            'demandas_carga_mental': 'DEMANDAS DE CARGA MENTAL',
            'demandas_emocionales': 'DEMANDAS EMOCIONALES',
            'exigencias_responsabilidad': 'EXIGENCIAS DE RESPONSABILIDAD',
            'consistencia_rol': 'CONSISTENCIA DE ROL',
            'demandas_ambientales': 'DEMANDAS AMBIENTALES Y ESFUERZO FÍSICO',
            'demandas_jornada': 'DEMANDAS DE JORNADA',
        }

        # Renderizar cada categoría
        for cat_key in ['demandas_cuantitativas', 'demandas_carga_mental', 'demandas_emocionales',
                        'exigencias_responsabilidad', 'consistencia_rol', 'demandas_ambientales', 'demandas_jornada']:
            if cat_key in categorias_dict:
                # Título de categoría
                ws.merge_cells(f'A{row}:I{row}')
                ws[f'A{row}'] = nombres_categorias.get(cat_key, cat_key.upper())
                ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
                ws[f'A{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                ws[f'A{row}'].border = border_thin
                row += 1

                # Items
                items = sorted(categorias_dict[cat_key], key=lambda x: x.item_numero)
                for eval in items:
                    ws.merge_cells(f'A{row}:F{row}')
                    ws[f'A{row}'] = f'  Item {eval.item_numero}'
                    ws[f'A{row}'].border = border_thin
                    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')

                    ws[f'G{row}'] = 'X' if eval.respuesta == 'si' else ''
                    ws[f'G{row}'].border = border_thin
                    ws[f'G{row}'].alignment = Alignment(horizontal='center')

                    ws[f'H{row}'] = 'X' if eval.respuesta == 'no' else ''
                    ws[f'H{row}'].border = border_thin
                    ws[f'H{row}'].alignment = Alignment(horizontal='center')

                    ws[f'I{row}'] = 'X' if eval.respuesta == 'na' else ''
                    ws[f'I{row}'].border = border_thin
                    ws[f'I{row}'].alignment = Alignment(horizontal='center')

                    row += 1

    row += 1

    # ===== SECCIÓN 6: CONCEPTO FINAL =====
    ws.merge_cells(f'A{row}:I{row}')
    cell = ws[f'A{row}']
    cell.value = '6. CONCEPTO FINAL'
    cell.font = seccion_font
    cell.fill = seccion_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    if concepto:
        campos_concepto = [
            ('Apariencia personal:', concepto.apariencia_personal),
            ('Actitud:', concepto.actitud),
            ('Concepto:', concepto.concepto),
            ('Recomendaciones:', concepto.recomendaciones),
        ]

        for label, valor in campos_concepto:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].fill = label_fill
            ws.merge_cells(f'B{row}:I{row}')
            ws[f'B{row}'] = valor or ''
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].border = border_thin
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 50
            row += 1

    row += 2

    # Firma
    if valoracion:
        ws[f'A{row}'] = 'Evaluador:'
        ws[f'A{row}'].font = label_font
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = valoracion.creado_por_usuario.nombre if hasattr(valoracion, 'creado_por_usuario') and valoracion.creado_por_usuario else ''
        ws[f'B{row}'].border = border_thin
        row += 1

        ws[f'A{row}'] = 'Fecha:'
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = valoracion.fecha_valoracion.strftime('%d/%m/%Y') if valoracion.fecha_valoracion else ''
        ws[f'B{row}'].border = border_thin

    # ===== GUARDAR O RETORNAR =====
    if guardar_archivo:
        # Generar nombre de archivo
        nombre = trabajador.nombre if trabajador and trabajador.nombre else "sin_nombre"
        documento = trabajador.documento if trabajador and trabajador.documento else "sin_documento"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        filename = f"valoracion_{sanitize_filename(nombre)}_{sanitize_filename(documento)}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Crear directorio si no existe
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        # Guardar
        wb.save(filepath)
        return filepath
    else:
        # Retornar BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


def calcular_edad(fecha_nacimiento: date) -> int:
    """Calcula la edad a partir de la fecha de nacimiento"""
    today = date.today()
    edad = today.year - fecha_nacimiento.year - (
        (today.month, today.day) < (fecha_nacimiento.month, fecha_nacimiento.day)
    )
    return edad


def generar_plantilla_vacia() -> BytesIO:
    """Genera un Excel vacío con solo el formato"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Valoración Ocupacional"

    ws['A1'] = 'VALORACIÓN OCUPACIONAL - PLANTILLA VACÍA'
    ws['A1'].font = Font(name='Arial', size=14, bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
