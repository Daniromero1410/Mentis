"""
Generador de Excel para valoraciones ocupacionales
Usa la plantilla plantilla_valoracion.xlsx y mapea los datos a las celdas específicas
"""
from openpyxl import load_workbook
from io import BytesIO
from typing import Optional, List
from datetime import date, datetime
import os
from pathlib import Path

# Ruta de la plantilla
PLANTILLA_PATH = os.path.join(os.path.dirname(__file__), 'plantilla_valoracion.xlsx')


def sanitize_filename(text: str) -> str:
    """Limpia un texto para usarlo como nombre de archivo"""
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        text = text.replace(char, '')
    text = text.replace(' ', '_')
    if len(text) > 50:
        text = text[:50]
    return text


def calcular_edad(fecha_nacimiento: date) -> int:
    """Calcula la edad a partir de la fecha de nacimiento"""
    today = date.today()
    edad = today.year - fecha_nacimiento.year - (
        (today.month, today.day) < (fecha_nacimiento.month, fecha_nacimiento.day)
    )
    return edad


def escribir_celda_segura(ws, celda: str, valor):
    """
    Escribe en una celda de forma segura, manejando celdas combinadas.
    Si la celda es parte de una merged cell, escribe en la celda superior izquierda.
    """
    try:
        # Intentar escribir directamente
        ws[celda] = valor
    except AttributeError:
        # Si falla, es una merged cell. Buscar la celda superior izquierda
        from openpyxl.utils import coordinate_to_tuple, get_column_letter

        for merged_range in ws.merged_cells.ranges:
            if celda in merged_range:
                # Obtener la celda superior izquierda del rango
                min_col = merged_range.min_col
                min_row = merged_range.min_row
                top_left = f"{get_column_letter(min_col)}{min_row}"
                ws[top_left] = valor
                return

        # Si no es una merged cell, re-lanzar el error
        raise


def generar_excel_valoracion(
    valoracion,
    trabajador,
    info_laboral,
    historia_ocupacional: List,
    actividad_laboral,
    evaluaciones: List,
    concepto,
    guardar_archivo: bool = True,
    output_dir: str = "app/services"
) -> Optional[BytesIO]:
    """
    Genera un archivo Excel de valoración usando la plantilla

    Returns:
        - Si guardar_archivo=True: ruta del archivo guardado
        - Si guardar_archivo=False: BytesIO con el contenido
    """

    # Cargar plantilla
    wb = load_workbook(PLANTILLA_PATH)
    ws = wb.active

    # ===== FECHA DE VALORACIÓN (Fila 3) =====
    if valoracion and valoracion.fecha_valoracion:
        escribir_celda_segura(ws, 'C3', valoracion.fecha_valoracion.day)
        escribir_celda_segura(ws, 'E3', valoracion.fecha_valoracion.month)
        escribir_celda_segura(ws, 'G3', valoracion.fecha_valoracion.year)

    # ===== IDENTIFICACIÓN =====
    if trabajador:
        # Nombre (B6)
        escribir_celda_segura(ws, 'B6', trabajador.nombre or '')

        # Documento (B7)
        escribir_celda_segura(ws, 'B7', trabajador.documento or '')

        # Identificación siniestro (B8)
        escribir_celda_segura(ws, 'B8', trabajador.identificacion_siniestro or '')

        # Fecha de nacimiento (C9, E9, G9) y edad (B10)
        if trabajador.fecha_nacimiento:
            escribir_celda_segura(ws, 'C9', trabajador.fecha_nacimiento.day)
            escribir_celda_segura(ws, 'E9', trabajador.fecha_nacimiento.month)
            escribir_celda_segura(ws, 'G9', trabajador.fecha_nacimiento.year)
            escribir_celda_segura(ws, 'B10', calcular_edad(trabajador.fecha_nacimiento))

        # Estado civil - marcar con X
        if trabajador.estado_civil:
            estado = trabajador.estado_civil.value if hasattr(trabajador.estado_civil, 'value') else str(trabajador.estado_civil)
            estado = estado.lower().replace(' ', '_')
            celdas_estado = {
                'casado': 'C11',
                'soltero': 'E11',
                'union_libre': 'G11',
                'separado': 'I11',
                'viudo': 'K11'
            }
            if estado in celdas_estado:
                escribir_celda_segura(ws, celdas_estado[estado], 'X')

        # Nivel educativo - marcar con X
        if trabajador.nivel_educativo:
            nivel = trabajador.nivel_educativo.lower()
            celdas_educacion = {
                'empirica': 'C12',
                'primaria': 'E12',
                'vocacional': 'G12',
                'bachillerato': 'I12',
                'tecnico': 'C13',
                'tecnologico': 'C13',
                'universitario': 'E13',
                'especializacion': 'G13',
                'postgrado': 'G13',
                'maestria': 'G13',
                'informal': 'I13',
                'analfabeta': 'K13'
            }
            for key, celda in celdas_educacion.items():
                if key in nivel:
                    escribir_celda_segura(ws, celda, 'X')
                    break

        # Formación específica (B14)
        escribir_celda_segura(ws, 'B14', trabajador.formacion_especifica or '')

        # Teléfonos (B15)
        escribir_celda_segura(ws, 'B15', trabajador.telefonos or '')

        # Dirección (B16)
        escribir_celda_segura(ws, 'B16', trabajador.direccion or '')

        # Zona - marcar con X
        if trabajador.zona:
            zona = trabajador.zona.value if hasattr(trabajador.zona, 'value') else str(trabajador.zona)
            if zona.lower() == 'urbano':
                escribir_celda_segura(ws, 'C17', 'X')
            elif zona.lower() == 'rural':
                escribir_celda_segura(ws, 'E17', 'X')

        # Diagnóstico mental (B18)
        escribir_celda_segura(ws, 'B18', trabajador.diagnostico_mental or '')

    # ===== INFORMACIÓN LABORAL =====
    if info_laboral:
        # Fecha evento AT/EL (B19)
        if info_laboral.fecha_evento_atel:
            escribir_celda_segura(ws, 'B19', info_laboral.fecha_evento_atel.strftime('%d/%m/%Y'))

        # Eventos no laborales
        if info_laboral.eventos_no_laborales:
            escribir_celda_segura(ws, 'C20', 'X')
            if info_laboral.evento_no_laboral_fecha:
                escribir_celda_segura(ws, 'G20', info_laboral.evento_no_laboral_fecha.strftime('%d/%m/%Y'))
            escribir_celda_segura(ws, 'I20', info_laboral.evento_no_laboral_diagnostico or '')
        else:
            escribir_celda_segura(ws, 'E20', 'X')

        # EPS (B21)
        escribir_celda_segura(ws, 'B21', info_laboral.eps or '')

        # Fondo pensión (B22)
        escribir_celda_segura(ws, 'B22', info_laboral.fondo_pension or '')

        # Días incapacidad (B23)
        escribir_celda_segura(ws, 'B23', info_laboral.dias_incapacidad or '')

        # Empresa (B26)
        escribir_celda_segura(ws, 'B26', info_laboral.empresa or '')

        # Vinculación laboral
        if info_laboral.vinculacion_laboral:
            escribir_celda_segura(ws, 'C27', 'X')
        else:
            escribir_celda_segura(ws, 'E27', 'X')

        # Tipo vinculación (B28)
        escribir_celda_segura(ws, 'B28', info_laboral.tipo_vinculacion or '')

        # Modalidad - marcar con X
        if info_laboral.modalidad:
            modalidad = info_laboral.modalidad.value if hasattr(info_laboral.modalidad, 'value') else str(info_laboral.modalidad)
            if 'presencial' in modalidad.lower():
                escribir_celda_segura(ws, 'C29', 'X')
            elif 'teletrabajo' in modalidad.lower():
                escribir_celda_segura(ws, 'E29', 'X')
            elif 'casa' in modalidad.lower():
                escribir_celda_segura(ws, 'G29', 'X')

        # Tiempo modalidad (B30)
        escribir_celda_segura(ws, 'B30', info_laboral.tiempo_modalidad or '')

        # NIT (B31)
        escribir_celda_segura(ws, 'B31', info_laboral.nit_empresa or '')

        # Fecha ingreso (C32, E32, G32)
        if info_laboral.fecha_ingreso_empresa:
            escribir_celda_segura(ws, 'C32', info_laboral.fecha_ingreso_empresa.day)
            escribir_celda_segura(ws, 'E32', info_laboral.fecha_ingreso_empresa.month)
            escribir_celda_segura(ws, 'G32', info_laboral.fecha_ingreso_empresa.year)

        # Antigüedad empresa (B33 años, D33 meses)
        escribir_celda_segura(ws, 'B33', info_laboral.antiguedad_empresa_anos or '')
        escribir_celda_segura(ws, 'D33', info_laboral.antiguedad_empresa_meses or '')

        # Antigüedad cargo (B34 años, D34 meses)
        escribir_celda_segura(ws, 'B34', info_laboral.antiguedad_cargo_anos or '')
        escribir_celda_segura(ws, 'D34', info_laboral.antiguedad_cargo_meses or '')

        # Contacto empresa (B35)
        escribir_celda_segura(ws, 'B35', info_laboral.contacto_empresa or '')

        # Correos (B36)
        escribir_celda_segura(ws, 'B36', info_laboral.correos or '')

        # Teléfonos empresa (B37)
        escribir_celda_segura(ws, 'B37', info_laboral.telefonos_empresa or '')

    # ===== HISTORIA OCUPACIONAL (Filas 42, 43, 44) =====
    filas_historia = [42, 43, 44]
    for i, hist in enumerate(historia_ocupacional[:3]):
        fila = filas_historia[i]
        escribir_celda_segura(ws, f'A{fila}', hist.empresa or '')
        escribir_celda_segura(ws, f'F{fila}', hist.cargo_funciones or '')
        escribir_celda_segura(ws, f'K{fila}', hist.duracion or '')
        escribir_celda_segura(ws, f'N{fila}', hist.motivo_retiro or '')

    # Otros oficios (B45)
    if actividad_laboral and actividad_laboral.otros_oficios:
        escribir_celda_segura(ws, 'B45', actividad_laboral.otros_oficios)

    # Oficios de interés (B46)
    if actividad_laboral and actividad_laboral.oficios_interes:
        escribir_celda_segura(ws, 'B46', actividad_laboral.oficios_interes)

    # ===== ACTIVIDAD LABORAL ACTUAL =====
    if actividad_laboral:
        # Nombre cargo (B49)
        escribir_celda_segura(ws, 'B49', actividad_laboral.nombre_cargo or '')

        # Tareas (B50)
        escribir_celda_segura(ws, 'B50', actividad_laboral.tareas or '')

        # Herramientas (B51)
        escribir_celda_segura(ws, 'B51', actividad_laboral.herramientas or '')

        # Horario (B52)
        escribir_celda_segura(ws, 'B52', actividad_laboral.horario or '')

        # EPP (B53)
        escribir_celda_segura(ws, 'B53', actividad_laboral.elementos_proteccion or '')

    # ===== FACTORES DE RIESGO PSICOSOCIALES =====
    # Mapeo de factores a filas según el mapa de celdas
    factores_filas = {
        # Demandas cuantitativas
        'ritmo_trabajo': 57, 'pausas': 58, 'tiempo_adicional': 59, 'volumen_carga': 60,
        # Carga mental
        'memoria_atencion': 62, 'detalle_precision': 63, 'info_presion': 64,
        'info_simultanea': 65, 'info_insuficiente': 66, 'carga_cognitiva': 67,
        'presion_tiempo': 68, 'agotamiento': 69,
        # Emocionales
        'trato_negativo': 71, 'situaciones_devastadoras': 72, 'impacto_extralaboral': 73,
        'errores': 74, 'tension': 75, 'monotonia': 76,
        # Responsabilidad
        'vida_salud': 78, 'supervision': 79, 'resultados': 80,
        'bienes': 81, 'dinero': 82, 'info_confidencial': 83,
        # Consistencia de rol
        'recursos': 85, 'ordenes_contradictorias': 86, 'requerimientos': 87,
        'principios_eticos': 88, 'variacion_tarea': 89, 'tareas_simultaneas': 90,
        'actualizacion': 91,
        # Ambientales
        'ruido': 93, 'iluminacion': 94, 'temperatura': 95, 'ventilacion': 96,
        'puesto': 97, 'orden': 98, 'biologicos': 99, 'quimicos': 100,
        'esfuerzo_fisico': 101, 'accidente': 102,
        # Jornada
        'nocturno': 104, 'dias_consecutivos': 105
    }

    # Columnas: M=alto, N=medio, O=bajo
    col_map = {'alto': 'M', 'medio': 'N', 'bajo': 'O'}

    for eval in evaluaciones:
        # El modelo usa 'calificacion' con valores: alto, medio, bajo
        if eval.calificacion:
            calificacion = eval.calificacion.value if hasattr(eval.calificacion, 'value') else str(eval.calificacion)
            calificacion = calificacion.lower()

            # Intentar mapear el item_texto a una fila
            item_texto_lower = eval.item_texto.lower() if eval.item_texto else ''

            # Mapeo de item_texto a clave de factor
            fila = None
            if 'ritmo' in item_texto_lower and 'acelerado' in item_texto_lower:
                fila = factores_filas.get('ritmo_trabajo')
            elif 'pausas' in item_texto_lower:
                fila = factores_filas.get('pausas')
            elif 'tiempo adicional' in item_texto_lower:
                fila = factores_filas.get('tiempo_adicional')
            elif 'volumen' in item_texto_lower and 'carga' in item_texto_lower:
                fila = factores_filas.get('volumen_carga')
            elif 'memoria' in item_texto_lower or 'atención' in item_texto_lower or 'concentración' in item_texto_lower:
                fila = factores_filas.get('memoria_atencion')
            elif 'detalle' in item_texto_lower or 'precisión' in item_texto_lower:
                fila = factores_filas.get('detalle_precision')
            elif 'información' in item_texto_lower and 'presión' in item_texto_lower:
                fila = factores_filas.get('info_presion')
            elif 'información' in item_texto_lower and 'simultánea' in item_texto_lower:
                fila = factores_filas.get('info_simultanea')
            elif 'información' in item_texto_lower and ('compleja' in item_texto_lower or 'insuficiente' in item_texto_lower):
                fila = factores_filas.get('info_insuficiente')
            elif 'carga cognitiva' in item_texto_lower:
                fila = factores_filas.get('carga_cognitiva')
            elif 'presión' in item_texto_lower and 'tiempo' in item_texto_lower:
                fila = factores_filas.get('presion_tiempo')
            elif 'agotamiento' in item_texto_lower:
                fila = factores_filas.get('agotamiento')
            elif 'trato negativo' in item_texto_lower or 'sentimientos' in item_texto_lower:
                fila = factores_filas.get('trato_negativo')
            elif 'devastadoras' in item_texto_lower:
                fila = factores_filas.get('situaciones_devastadoras')
            elif 'extralaboral' in item_texto_lower:
                fila = factores_filas.get('impacto_extralaboral')
            elif 'errores' in item_texto_lower:
                fila = factores_filas.get('errores')
            elif 'tensión' in item_texto_lower:
                fila = factores_filas.get('tension')
            elif 'monotonía' in item_texto_lower or 'repetitiva' in item_texto_lower:
                fila = factores_filas.get('monotonia')
            elif 'vida' in item_texto_lower or 'salud' in item_texto_lower:
                fila = factores_filas.get('vida_salud')
            elif 'supervisión' in item_texto_lower:
                fila = factores_filas.get('supervision')
            elif 'resultados' in item_texto_lower:
                fila = factores_filas.get('resultados')
            elif 'bienes' in item_texto_lower:
                fila = factores_filas.get('bienes')
            elif 'dinero' in item_texto_lower:
                fila = factores_filas.get('dinero')
            elif 'confidencial' in item_texto_lower:
                fila = factores_filas.get('info_confidencial')
            elif 'recursos' in item_texto_lower or 'herramientas' in item_texto_lower:
                fila = factores_filas.get('recursos')
            elif 'contradictorias' in item_texto_lower:
                fila = factores_filas.get('ordenes_contradictorias')
            elif 'innecesarios' in item_texto_lower:
                fila = factores_filas.get('requerimientos')
            elif 'éticos' in item_texto_lower or 'ética' in item_texto_lower:
                fila = factores_filas.get('principios_eticos')
            elif 'variación' in item_texto_lower:
                fila = factores_filas.get('variacion_tarea')
            elif 'simultáneas' in item_texto_lower and 'tareas' in item_texto_lower:
                fila = factores_filas.get('tareas_simultaneas')
            elif 'actualización' in item_texto_lower:
                fila = factores_filas.get('actualizacion')
            elif 'ruido' in item_texto_lower:
                fila = factores_filas.get('ruido')
            elif 'iluminación' in item_texto_lower:
                fila = factores_filas.get('iluminacion')
            elif 'temperatura' in item_texto_lower:
                fila = factores_filas.get('temperatura')
            elif 'ventilación' in item_texto_lower:
                fila = factores_filas.get('ventilacion')
            elif 'puesto' in item_texto_lower or 'distribución' in item_texto_lower:
                fila = factores_filas.get('puesto')
            elif 'orden' in item_texto_lower or 'aseo' in item_texto_lower:
                fila = factores_filas.get('orden')
            elif 'biológicos' in item_texto_lower:
                fila = factores_filas.get('biologicos')
            elif 'químicos' in item_texto_lower:
                fila = factores_filas.get('quimicos')
            elif 'esfuerzo' in item_texto_lower and 'físico' in item_texto_lower:
                fila = factores_filas.get('esfuerzo_fisico')
            elif 'accidente' in item_texto_lower:
                fila = factores_filas.get('accidente')
            elif 'nocturno' in item_texto_lower:
                fila = factores_filas.get('nocturno')
            elif 'consecutivo' in item_texto_lower:
                fila = factores_filas.get('dias_consecutivos')

            # Si encontramos la fila y la calificación es válida, marcar con X
            if fila and calificacion in col_map:
                col = col_map[calificacion]
                escribir_celda_segura(ws, f'{col}{fila}', 'X')

    # ===== CONCEPTO FINAL =====
    if concepto:
        # Concepto psicológico (A108)
        # El concepto puede estar en concepto_generado o concepto_editado
        concepto_texto = concepto.concepto_editado or concepto.concepto_generado or ''
        escribir_celda_segura(ws, 'A108', concepto_texto)

        # Orientación reintegro (A115)
        escribir_celda_segura(ws, 'A115', concepto.orientacion_reintegro or '')

        # Firmas
        # Elaboró (A124)
        escribir_celda_segura(ws, 'A124', concepto.elaboro_nombre or '')
        
        # Insertar imagen de firma Elaboró si existe
        if concepto.elaboro_firma and os.path.exists(concepto.elaboro_firma):
            try:
                from openpyxl.drawing.image import Image as ExcelImage
                img = ExcelImage(concepto.elaboro_firma)
                # Ajustar tamaño si es necesario (ej: 150px ancho)
                aspect_ratio = img.height / img.width
                img.width = 150
                img.height = 150 * aspect_ratio
                
                # Posicionar sobre A122 (un poco arriba del nombre)
                ws.add_image(img, 'A122')
            except Exception as e:
                print(f"Error insertando firma elaboró: {e}")

        # Revisión (H124)
        escribir_celda_segura(ws, 'H124', concepto.reviso_nombre or '')
        
        # Insertar imagen de firma Revisó si existe
        if concepto.reviso_firma and os.path.exists(concepto.reviso_firma):
            try:
                from openpyxl.drawing.image import Image as ExcelImage
                img = ExcelImage(concepto.reviso_firma)
                # Ajustar tamaño
                aspect_ratio = img.height / img.width
                img.width = 150
                img.height = 150 * aspect_ratio
                
                # Posicionar sobre H122
                ws.add_image(img, 'H122')
            except Exception as e:
                print(f"Error insertando firma revisó: {e}")

    # ===== GUARDAR O RETORNAR =====
    if guardar_archivo:
        # Generar nombre de archivo
        nombre = trabajador.nombre if trabajador and trabajador.nombre else "sin_nombre"
        documento = trabajador.documento if trabajador and trabajador.documento else "sin_documento"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        filename = f"valoracion_{sanitize_filename(nombre)}_{sanitize_filename(documento)}_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Crear directorio si no existe
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        # Guardar
        wb.save(filepath)
        return filepath
    else:
        # Retornar BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output


def generar_plantilla_vacia() -> BytesIO:
    """Genera una copia de la plantilla vacía"""
    wb = load_workbook(PLANTILLA_PATH)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
